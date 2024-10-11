import requests
from bs4 import BeautifulSoup
url = "https://docs.google.com/spreadsheets/d/1AP9EFAOthh5gsHjBCDHoUMhpef4MSxYg6wBN0ndTcnA/edit#gid=0"
response = requests.get(url)
html_content = response.content
soup = BeautifulSoup(html_content, "html.parser")
first_cell = soup.find("td", {"class": "s2"}).text.strip()
if first_cell != "Aktif":
    exit()
first_cell = soup.find("td", {"class": "s1"}).text.strip()
print(first_cell)

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from io import BytesIO
import numpy as np
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import time
from tqdm import tqdm
import warnings
from colorama import init, Fore, Style
import threading
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import chromedriver_autoinstaller
from concurrent.futures import ThreadPoolExecutor
import subprocess
from selenium.common.exceptions import TimeoutException, WebDriverException
from bs4 import BeautifulSoup
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from datetime import datetime
from datetime import datetime, timedelta
import shutil
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from pathlib import Path
import re
import http.client
import json
import gc
from concurrent.futures import ThreadPoolExecutor, as_completed
warnings.filterwarnings("ignore")
pd.options.mode.chained_assignment = None


print("Oturum Açma Başarılı Oldu")
print(" /﹋\ ")
print("(҂`_´)")
print(Fore.RED + "<,︻╦╤─ ҉ - -")
print("/﹋\\")
print("Mustafa ARI")
print(" ")
print(Fore.RED + "Zamanlı Kategori Sıralamaları")




#region İÇ GİYİM Ürünlerini İç Giyim Tüm Ürünler Kategorisine Alma ve Firma Bazlı Sıralama (İlk Ürünler Emin Abinin Belirlediği Ürünler)

# ChromeOptions oluştur
chromedriver_autoinstaller.install()
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--log-level=1') 
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])  
driver = webdriver.Chrome(options=chrome_options)

login_url = "https://task.haydigiy.com/kullanici-giris/?ReturnUrl=%2Fadmin"
driver.get(login_url)

email_input = driver.find_element("id", "EmailOrPhone")
email_input.send_keys("mustafa_kod@haydigiy.com")

password_input = driver.find_element("id", "Password")
password_input.send_keys("123456")
password_input.send_keys(Keys.RETURN)

# Giriş yaptıktan sonra belirtilen sayfaya git
desired_page_url = "https://task.haydigiy.com/admin/product/bulkedit/"
driver.get(desired_page_url)

#Kategori Dahil Alan
category_select = Select(driver.find_element("id", "SearchInCategoryIds"))

#Kategori ID'si (Fiyata Hamle)
category_select.select_by_value("172")

#Seçiniz Kısmının Tikini Kaldırma
all_remove_buttons = driver.find_elements(By.XPATH, "//span[@class='select2-selection__choice__remove']")
if len(all_remove_buttons) > 1:
    second_remove_button = all_remove_buttons[1]
    second_remove_button.click()
else:
    pass

#Ara Butonuna Tıklama
search_button = driver.find_element(By.ID, "search-products")
search_button.click()

# Sayfanın en sonuna git
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

# Kategori Güncelleme Tikine Tıklama
checkbox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "Category_Update")))
driver.execute_script("arguments[0].click();", checkbox)

# Kategori Güncelleme Alanını Bulma
category_id_select = driver.find_element(By.ID, "CategoryId")

# Kategori Güncelleme Alanında Kategori ID'si Seçme (Fiyata Hamle)
category_id_select = Select(category_id_select)
category_id_select.select_by_value("264")

# Kategori Güncelleme Alanında Yapılacak İşlem Alanını Bulma
category_transaction_select = driver.find_element(By.ID, "CategoryTransactionId")

# Kategori Güncelleme Alanında Yapılacak İşlemi Seçme (Kategoriden Çıkar)
category_transaction_select = Select(category_transaction_select)
category_transaction_select.select_by_value("0")

try:
    #Kaydet Butonunu Bulma
    save_button = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, 'btn-primary'))
    )

    #Kaydet Butonununa Tıklama
    driver.execute_script("arguments[0].click();", save_button)
except Exception as e:
    print(f"Hata: {e}")

# Tarayıcıyı kapat
driver.quit()







# İstek gönder
url = "https://task.haydigiy.com/FaprikaXls/MGMMC3/1/"
response = requests.get(url)

# Excel dosyasını indir
with open("veriler.xlsx", "wb") as file:
    file.write(response.content)

# Excel dosyasını yükle
veri = pd.read_excel("veriler.xlsx")

# "UrunAdi" ve "Kategori" dışındaki sütunları sil
sutunlar_sil = [sutun for sutun in veri.columns if sutun not in ["UrunAdi", "Kategori"]]
veri.drop(columns=sutunlar_sil, inplace=True)

# Tüm tablodaki yenilenen değerleri teke düşür (benzersiz yap)
veri.drop_duplicates(inplace=True)

# Excel dosyasını yeniden kaydet
veri.to_excel("veriler.xlsx", index=False)




# Excel dosyasını yükle
veri = pd.read_excel("veriler.xlsx")

# "UrunAdi" sütununda ".1364." içeren hücreleri içeren satırları al
urun_1364 = veri[veri["UrunAdi"].str.contains(".1364.")]

# ".1364." içeren satırları içeren Excel dosyasını kaydet
urun_1364.drop(columns=["Kategori"], inplace=True)
urun_1364.to_excel("urun_1364.xlsx", index=False)

# ".1783." içeren hücreleri içeren satırları al
urun_1783 = veri[veri["UrunAdi"].str.contains(".1783.")]

# ".1783." içeren satırları içeren Excel dosyasını kaydet
urun_1783.drop(columns=["Kategori"], inplace=True)
urun_1783.to_excel("urun_1783.xlsx", index=False)

# ".1231." içeren hücreleri içeren satırları al
urun_1231 = veri[veri["UrunAdi"].str.contains(".1231.")]

# ".1231." içeren satırları içeren Excel dosyasını kaydet
urun_1231.drop(columns=["Kategori"], inplace=True)
urun_1231.to_excel("urun_1231.xlsx", index=False)

# "Kategori" sütununda "İfondi Sıralama ilk 10 Ürün" içeren hücreleri içeren satırları al
kategori_ifondi = veri[veri["Kategori"].str.contains("İfondi Sıralama ilk 10 Ürün")]

# "İfondi Sıralama ilk 10 Ürün" içeren satırları içeren Excel dosyasını kaydet
kategori_ifondi.drop(columns=["Kategori"], inplace=True)
kategori_ifondi.to_excel("kategori_ifondi.xlsx", index=False)

# Son olarak veriler excelini silelim
import os
os.remove("veriler.xlsx")




# Kategori ifondi excelini yükle
kategori_ifondi = pd.read_excel("kategori_ifondi.xlsx")

# Diğer üç excel dosyasını yükle
urun_1231 = pd.read_excel("urun_1231.xlsx")
urun_1364 = pd.read_excel("urun_1364.xlsx")
urun_1783 = pd.read_excel("urun_1783.xlsx")

# Başlangıç indeksleri
index_1231 = 0
index_1364 = 0
index_1783 = 0

# Kategori ifondi exceline ekleme döngüsü
while True:
    # Urun_1231'den veri ekle
    if index_1231 < len(urun_1231):
        kategori_ifondi = pd.concat([kategori_ifondi, pd.DataFrame({"UrunAdi": [urun_1231.iloc[index_1231]["UrunAdi"]]})], ignore_index=True)
        index_1231 += 1
    
    # Urun_1364'ten veri ekle
    if index_1364 < len(urun_1364):
        kategori_ifondi = pd.concat([kategori_ifondi, pd.DataFrame({"UrunAdi": [urun_1364.iloc[index_1364]["UrunAdi"]]})], ignore_index=True)
        index_1364 += 1
    
    # Urun_1783'ten veri ekle
    if index_1783 < len(urun_1783):
        kategori_ifondi = pd.concat([kategori_ifondi, pd.DataFrame({"UrunAdi": [urun_1783.iloc[index_1783]["UrunAdi"]]})], ignore_index=True)
        index_1783 += 1
    
    # Tüm veriler eklenene kadar devam et
    if index_1231 >= len(urun_1231) and index_1364 >= len(urun_1364) and index_1783 >= len(urun_1783):
        break


# Sonucu yeni bir Excel dosyasına kaydet
kategori_ifondi.to_excel("kategori_ifondi_yeni.xlsx", index=False)


# Kategori ifondi excelini yükle
veri = pd.read_excel("kategori_ifondi_yeni.xlsx")

# Tüm tablodaki yenilenen değerleri teke düşür (benzersiz yap)
veri.drop_duplicates(inplace=True)

# Excel dosyasını yeniden kaydet
veri.to_excel("kategori_ifondi_yeni.xlsx", index=False)


# Excel dosyalarını sil
os.remove("kategori_ifondi.xlsx")
os.remove("urun_1231.xlsx")
os.remove("urun_1364.xlsx")
os.remove("urun_1783.xlsx")



# XML'den Ürün Bilgilerini Çekme ve Temizleme
xml_url = "https://task.haydigiy.com/FaprikaXml/1N8OKV/1/"
response = requests.get(xml_url)
xml_data = response.text
soup = BeautifulSoup(xml_data, 'xml')

product_data = []
for item in soup.find_all('item'):
    title = item.find('title').text
    title_cleaned = re.sub(r' - H.*', '', title)
    product_id = item.find('g:id').text if item.find('g:id') else None
    product_data.append({'UrunAdi': title_cleaned, 'ID': product_id})

df_xml = pd.DataFrame(product_data)

# Excel ile Birleştirme
df_calisma_alani = pd.read_excel('kategori_ifondi_yeni.xlsx')
df_merged = pd.merge(df_calisma_alani, df_xml, how='left', left_on='UrunAdi', right_on='UrunAdi')

# Sonuçları Mevcut Excel Dosyasının Üzerine Kaydetme
df_merged.to_excel('kategori_ifondi_yeni.xlsx', index=False)



# Kategori ifondi_yeni.xlsx dosyasını yükle
kategori_ifondi_yeni = pd.read_excel("kategori_ifondi_yeni.xlsx")

# "ID" sütununda boş olan satırları sil
kategori_ifondi_yeni.dropna(subset=["ID"], inplace=True)

# "UrunAdi" sütununu sil
kategori_ifondi_yeni.drop(columns=["UrunAdi"], inplace=True)

# Temizlenmiş veriyi aynı Excel dosyasının üzerine kaydet
kategori_ifondi_yeni.to_excel("kategori_ifondi_yeni.xlsx", index=False)






# "Birlestirilmis_Veriler_Sirali_Yeni" Excel dosyasını oku
df_birlestirilmis_sirali_yeni = pd.read_excel('kategori_ifondi_yeni.xlsx', sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter('kategori_ifondi_yeni.xlsx', engine='openpyxl') as writer:
    # Her bir sayfa için işlemleri gerçekleştir
    for sheet_name, df_sheet_birlestirilmis_sirali_yeni in df_birlestirilmis_sirali_yeni.items():
        # Her satıra sırasıyla -50, -49, -48, ... şeklinde sayıları ekle
        df_sheet_birlestirilmis_sirali_yeni['Numara'] = range(-len(df_sheet_birlestirilmis_sirali_yeni), 0)

        # Veriyi güncellenmiş haliyle aynı sayfaya kaydet
        df_sheet_birlestirilmis_sirali_yeni.to_excel(writer, sheet_name=sheet_name, index=False)




# Kategori ifondi_yeni.xlsx dosyasını yükle
kategori_ifondi_yeni = pd.read_excel("kategori_ifondi_yeni.xlsx")

# "SayfaIsmi" adında yeni bir sütun oluştur ve tüm hücreleri "İç Giyim Tüm Ürünler" olarak doldur
kategori_ifondi_yeni["Kategori ID"] = 264

# Veriyi aynı Excel dosyasının üzerine kaydet
kategori_ifondi_yeni.to_excel("kategori_ifondi_yeni.xlsx", index=False)

# Yeni dosya adı
yeni_dosya_adı = "Kategori Sıralama.xlsx"

# Eski dosyanın adını değiştir
os.rename("kategori_ifondi_yeni.xlsx", yeni_dosya_adı)




# ChromeOptions oluştur
chromedriver_autoinstaller.install()
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--log-level=1') 
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])  
driver = webdriver.Chrome(options=chrome_options)


# Tarayıcı penceresini belirli bir boyuta getirme
driver.set_window_size(50, 50)


# Giriş
login_url = "https://task.haydigiy.com/kullanici-giris/?ReturnUrl=%2Fadmin"

# Verilen URL'ye gitme ve giriş yapma fonksiyonu
def login(username, password):
    driver.get(login_url)
    email_input = driver.find_element("id", "EmailOrPhone")
    email_input.send_keys(username)
    password_input = driver.find_element("id", "Password")
    password_input.send_keys(password)
    password_input.send_keys(Keys.ENTER)
    time.sleep(2)

# order_edit_urls listesi
order_edit_urls = [
        "https://task.haydigiy.com/Admin/Category/Sort/264"

]

try:
    # Giriş yapma
    login("mustafa_kod@haydigiy.com", "123456")

    # Her bir link için işlem yapma
    for url in order_edit_urls:
        driver.get(url)

        js_code = """
        document.getElementById("SortOptionId").value = "25";
        var event = new Event('change');
        document.getElementById("SortOptionId").dispatchEvent(event);

        document.getElementById("btnChangeSorting").click();
        document.getElementById("btnChangeSorting-action-confirmation-submit-button").click();
        """
        driver.execute_script(js_code)

        # İşlem tamamlanana kadar bekleyin (maksimum 10 saniye)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "btnChangeSorting")))

except Exception as e:
    print("Hata oluştu:", e)

# Tarayıcıyı kapatma
driver.quit()



# Global tken değişkeni
_auth_token = None

# Token alma fonksiyonu
def get_auth_token():
    global _auth_token
    if _auth_token is None:  
        login_url = "https://siparis.haydigiy.com/api/customer/login"
        login_payload = {
            "apiKey": "MypGcaEInEOTzuYQydgDHQ",
            "secretKey": "jRqliBLDPke76YhL_WL5qg",
            "emailOrPhone": "mustafa_kod@haydigiy.com",
            "password": "123456"
        }
        login_headers = {
            "Content-Type": "application/json"
        }

        response = requests.post(login_url, json=login_payload, headers=login_headers)
        if response.status_code == 200:
            _auth_token = response.json().get("data", {}).get("token")
            if not _auth_token:
                raise Exception("TOKEN ALINAMADI")
        else:
            raise Exception(f"GİRİŞ BAŞARISIZ: {response.text}")
    return _auth_token

# Token alma işlemi
token = get_auth_token()
df = pd.read_excel("Kategori Sıralama.xlsx")
conn = http.client.HTTPSConnection("siparis.haydigiy.com")


for index, row in tqdm(df.iterrows(), total=df.shape[0], desc="İç Giyim Kategorisi Sıralanıyor"):

    category_id = row['Kategori ID']
    display_order = row['Numara']
    product_id = row['ID']


    product_id = str(row['ID']).replace(".0", "")

    payload = json.dumps({
        "CategoryId": int(category_id),  
        "IsFeaturedProduct": False, 
        "DisplayOrder": int(display_order)
    })

    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {token}',
        'Cookie': '.Application.Customer=64684894-1b54-488d-bd59-76b94842df65'
    }


    conn.request("PUT", f"/adminapi/product/product-categories?productId={product_id}", payload, headers)
    res = conn.getresponse()
    data = res.read()

conn.close()



# Silmek istediğiniz dosyaların listesi
dosya_listesi = ['Kategori Sıralama.xlsx']

# Dosyaları silme işlemi
for dosya in dosya_listesi:
    try:
        os.remove(dosya)
    
    except FileNotFoundError:
        print(f"'{dosya}' dosyası bulunamadı.")
    except Exception as e:
        print(f"'{dosya}' dosyasını silerken bir hata oluştu: {str(e)}")

#endregion

#region Tesettür Kategorisi Sıralama

# ChromeOptions oluştur
chromedriver_autoinstaller.install()
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--log-level=1') 
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])  
driver = webdriver.Chrome(options=chrome_options)


# Tarayıcı penceresini belirli bir boyuta getirme
driver.set_window_size(50, 50)


# Giriş
login_url = "https://task.haydigiy.com/kullanici-giris/?ReturnUrl=%2Fadmin"

# Verilen URL'ye gitme ve giriş yapma fonksiyonu
def login(username, password):
    driver.get(login_url)
    email_input = driver.find_element("id", "EmailOrPhone")
    email_input.send_keys(username)
    password_input = driver.find_element("id", "Password")
    password_input.send_keys(password)
    password_input.send_keys(Keys.ENTER)
    time.sleep(2)

# order_edit_urls listesi
order_edit_urls = [
        "https://task.haydigiy.com/Admin/Category/Sort/502"

]

try:
    # Giriş yapma
    login("mustafa_kod@haydigiy.com", "123456")

    # Her bir link için işlem yapma
    for url in order_edit_urls:
        driver.get(url)

        js_code = """
        document.getElementById("SortOptionId").value = "25";
        var event = new Event('change');
        document.getElementById("SortOptionId").dispatchEvent(event);

        document.getElementById("btnChangeSorting").click();
        document.getElementById("btnChangeSorting-action-confirmation-submit-button").click();
        """
        driver.execute_script(js_code)

        # İşlem tamamlanana kadar bekleyin (maksimum 10 saniye)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "btnChangeSorting")))

except Exception as e:
    print("Hata oluştu:", e)

# Tarayıcıyı kapatma
driver.quit()






def get_excel_data(url):
    response = requests.get(url)

    if response.status_code == 200:
        # Excel dosyasını oku
        df = pd.read_excel(BytesIO(response.content))
        return df
    else:
        
        return None

# İlk linkten veriyi al
url1 = "https://task.haydigiy.com/FaprikaXls/5DO7BZ/1/"
data1 = get_excel_data(url1)

# İkinci linkten veriyi al
url2 = "https://task.haydigiy.com/FaprikaXls/5DO7BZ/2/"
data2 = get_excel_data(url2)

# İki veriyi birleştir
if data1 is not None and data2 is not None:
    merged_data = pd.concat([data1, data2], ignore_index=True)

    # Gereksiz sütunları sil
    columns_to_keep = ["StokAdedi", "UrunAdi", "AlisFiyati", "SatisFiyati", "Kategori", "MorhipoKodu", "HepsiBuradaKodu"]
    merged_data = merged_data[columns_to_keep]

    # Birleştirilmiş veriyi Excel dosyasına kaydet
    merged_data.to_excel("birlesmis__veri.xlsx", index=False)

    # Birleştirilmiş veriyi oku
    final_data = pd.read_excel("birlesmis__veri.xlsx")

else:
    pass


# Veriyi Okuma
df = pd.read_excel('birlesmis__veri.xlsx')

# Alış Fiyatına Göre İşlemler ve Kategori Kontrolü
def calculate_list_price(row):
    alis_fiyati = row['AlisFiyati']
    kategori = row['Kategori']

    if 0 <= alis_fiyati <= 24.99:
        result = alis_fiyati + 10
    elif 25 <= alis_fiyati <= 39.99:
        result = alis_fiyati + 13
    elif 40 <= alis_fiyati <= 59.99:
        result = alis_fiyati + 17
    elif 60 <= alis_fiyati <= 200.99:
        result = alis_fiyati * 1.30
    elif alis_fiyati >= 201:
        result = alis_fiyati * 1.25
    else:
        result = alis_fiyati

    # KDV
    if isinstance(kategori, str) and any(category in kategori for category in ["Parfüm", "Gözlük", "Saat"]):
        result *= 1.20
    else:
        result *= 1.10

    return result


# Yeni Sütun Oluşturma
df['ListeFiyati'] = df.apply(calculate_list_price, axis=1)

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('birlesmis__veri.xlsx', index=False)











# Veriyi Okuma
df = pd.read_excel('birlesmis__veri.xlsx')

# "StokAdedi" Sütununda 0'dan Büyük Olan Değerlerin Adedi
df['StokAdedi_GT_0'] = df['StokAdedi'].apply(lambda x: 1 if x > 0 else 0)
stok_adedi_gt_0_adet = df.groupby('UrunAdi')['StokAdedi_GT_0'].sum().reset_index()

# "UrunAdi" Sütunundaki Toplam Yenilenme Adedi
toplam_yenilenme_adedi = df.groupby('UrunAdi').size().reset_index(name='ToplamYenilenmeAdedi')

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0'], axis=1, errors='ignore')


# Oranı Hesapla ve Yeni Sütunu Ekle
df = pd.merge(df, stok_adedi_gt_0_adet, on='UrunAdi', how='left')
df = pd.merge(df, toplam_yenilenme_adedi, on='UrunAdi', how='left')
df['Beden Durumu'] = df['StokAdedi_GT_0'] / df['ToplamYenilenmeAdedi']

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0', 'ToplamYenilenmeAdedi'], axis=1, errors='ignore')

# Oranı 100 ile Çarpma
df['Beden Durumu'] *= 100

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('birlesmis__veri.xlsx', index=False)











# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# "ListeFiyati" ve "SatisFiyati" sütunlarındaki verilerden işlem yap
birlesmis_veri["İndirimOrani"] = (birlesmis_veri["ListeFiyati"] - birlesmis_veri["SatisFiyati"]) * 100 / birlesmis_veri["ListeFiyati"]

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)








# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# İstenmeyen sütunları sil
silinecek_sutunlar = ["StokAdedi", "AlisFiyati", "SatisFiyati", "ListeFiyati"]
birlesmis_veri = birlesmis_veri.drop(columns=silinecek_sutunlar, errors='ignore')

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)







# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# Tüm tablodaki benzersiz değerleri teke düşür
birlesmis_veri = birlesmis_veri.drop_duplicates()

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)






# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# "Beden Durumu" sütunundaki değeri 50'den küçük olan satırları filtrele
birlesmis_veri = birlesmis_veri[birlesmis_veri["Beden Durumu"] >= 50]

# İstenmeyen sütunları sil
silinecek_sutunlar = ["Beden Durumu"]
birlesmis_veri = birlesmis_veri.drop(columns=silinecek_sutunlar, errors='ignore')

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)







# XML'den Ürün Bilgilerini Çekme ve Temizleme
xml_url = "https://task.haydigiy.com/FaprikaXml/1YJGSU/1/"
response = requests.get(xml_url)
xml_data = response.text
soup = BeautifulSoup(xml_data, 'xml')

product_data = []
for item in soup.find_all('item'):
    title = item.find('title').text.replace(' - Haydigiy', '')
    product_id = item.find('g:id').text if item.find('g:id') else None
    product_data.append({'UrunAdi': title, 'ID': product_id})

df_xml = pd.DataFrame(product_data)


# Excel ile Birleştirme
df_calisma_alani = pd.read_excel('birlesmis__veri.xlsx')
df_merged = pd.merge(df_calisma_alani, df_xml, how='left', left_on='UrunAdi', right_on='UrunAdi')

# Sonuçları Mevcut Excel Dosyasının Üzerine Kaydetme
df_merged.to_excel('birlesmis__veri.xlsx', index=False)






# birlesmis_veri Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx', sheet_name='Sheet1')

# "ID" sütunundaki boş olan hücreleri içeren satırları filtrele
df_birlesmis_veri = df_birlesmis_veri.dropna(subset=['ID'])

# Sonuçları güncellenmiş haliyle aynı Excel dosyasına kaydet
df_birlesmis_veri.to_excel('birlesmis__veri.xlsx', sheet_name='Sheet1', index=False)





# Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx')

# 'UrunAdi' sütununu sil
df_birlesmis_veri = df_birlesmis_veri.drop(columns=['UrunAdi'], errors='ignore')

# 'ID' sütununu en başa al
df_birlesmis_veri = df_birlesmis_veri[['ID'] + [col for col in df_birlesmis_veri.columns if col != 'ID']]

# Veriyi Excel dosyasına kaydet (üzerine yaz)
df_birlesmis_veri.to_excel('birlesmis__veri.xlsx', index=False)









# Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx')

# 'Kategori' sütununda "Tekrar Stokta" içeren hücreleri filtrele
df_birlesmis_veri['Tekrar Stokta'] = df_birlesmis_veri['Kategori'].apply(lambda x: 'Tekrar Stokta' if 'Tekrar Stokta' in str(x) else None)

# Veriyi Excel dosyasına kaydet (üzerine yaz)
df_birlesmis_veri.to_excel('birlesmis__veri.xlsx', index=False)








# Excel dosyasını oku
df = pd.read_excel('birlesmis__veri.xlsx')

# 'MorhipoKodu' sütunundaki değerleri sayıya dönüştür ve boş olan hücrelere "0" yaz
df['MorhipoKodu'] = pd.to_numeric(df['MorhipoKodu'], errors='coerce').fillna(0).astype(int)

# 'MorhipoKodu' sütunundaki değerleri sayıya dönüştür ve boş olan hücrelere "0" yaz
df['HepsiBuradaKodu'] = pd.to_numeric(df['HepsiBuradaKodu'], errors='coerce').fillna(0).astype(int)

# Güncellenmiş DataFrame'i aynı Excel dosyasının üzerine yaz
df.to_excel('birlesmis__veri.xlsx', index=False)











# Excel dosyasını oku
df = pd.read_excel('birlesmis__veri.xlsx')

# 'BaşariOrani' sütununu oluştur ve işlemi gerçekleştir
df['BaşariOrani'] = df['MorhipoKodu'] / df['HepsiBuradaKodu'].replace(0, 1)

# 'HepsiBuradaKodu' sütunundaki 0 değerlerini 'BaşariOrani' sütununa 0 olarak yaz
df.loc[df['HepsiBuradaKodu'] == 0, 'BaşariOrani'] = 0

# 'İndirimOrani' sütunundaki değerlerde 1'den küçük olanları "0" ile değiştir
df.loc[df['İndirimOrani'] < 1, 'İndirimOrani'] = 0

# Güncellenmiş DataFrame'i aynı Excel dosyasının üzerine yaz
df.to_excel('birlesmis__veri.xlsx', index=False)







# Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx')

# İstenilen kategorileri içeren satırları filtrele
istenen_kategoriler = [
    'TESETTÜR'
]

for kategori in istenen_kategoriler:
    df_kategori = df_birlesmis_veri[df_birlesmis_veri['Kategori'].str.contains(kategori, case=False, na=False)]
    
    # Veriyi yeni bir sayfaya kaydet (excel dosyasının içine eklenir)
    with pd.ExcelWriter('birlesmis__veri.xlsx', engine='openpyxl', mode='a') as writer:
        df_kategori.to_excel(writer, sheet_name=f'{kategori}', index=False)






# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx", sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter("birlesmis__veri.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    # Her sayfa için işlemi yap
    for sheet_name, df in birlesmis_veri.items():
        # Yeni sütunu hesapla
        df["Görüntülenme"] = (100 *  df["HepsiBuradaKodu"] / df["HepsiBuradaKodu"].sum())

        # Sonucu Excel dosyasına yaz
        df.to_excel(writer, sheet_name=sheet_name, index=False)






# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx", sheet_name=None)

# Her sayfa için işlemi yap
for sheet_name, df in birlesmis_veri.items():
    # "Görüntülenme" sütununu sayısal değerlere dönüştür
    df["Görüntülenme"] = pd.to_numeric(df["Görüntülenme"], errors='coerce')
    
    # Yeni sütunu hesapla
    ortalama = 0.7 * df["Görüntülenme"].mean()
    
    # Değerleri güncelle
    df["Görüntülenme"] = np.where(df["Görüntülenme"] < ortalama, "Az Görüntülenme", "")


    df.drop(columns=["HepsiBuradaKodu"], inplace=True)

    # Excel dosyasına yaz
    with pd.ExcelWriter("birlesmis__veri.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)







# Yeni bir Excel dosyası oluştur
yeni_birlesmis_veri = pd.ExcelWriter("yeni_birlesmis__veri.xlsx")

# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx", sheet_name=None)

# Her sayfa için işlemi yap
for sheet_name, df in birlesmis_veri.items():
    # Eğer veri yoksa işlem yapma
    if df.empty:
        continue

    # Yeni bir veri çerçevesi oluştur
    yeni_df = pd.DataFrame()

    # MorhipoKodu sütunundaki en büyük ilk 6 değeri al ve Sonuç sütununa ekle
    morhipo_kodu_top_6 = df.nlargest(6, "MorhipoKodu")
    for index, row in morhipo_kodu_top_6.iterrows():
        yeni_df.loc[len(yeni_df), "Sonuç"] = row["ID"]
    df.drop(morhipo_kodu_top_6.index, inplace=True, errors="ignore")

    while not df.empty:
        # Tekrar Stokta sütununda "Tekrar Stokta" verisi var mı kontrol et
        if "Tekrar Stokta" in df.columns and "Tekrar Stokta" in df["Tekrar Stokta"].unique():
            # Tekrar Stokta sütunundaki ilk "Tekrar Stokta" değerini al ve Sonuç sütununa ekle
            try:
                tekrar_stokta_first = df[df["Tekrar Stokta"] == "Tekrar Stokta"]["ID"].iloc[0]
                yeni_df.loc[len(yeni_df), "Sonuç"] = tekrar_stokta_first
                df.drop(df[df["ID"] == tekrar_stokta_first].index, inplace=True, errors="ignore")
            except IndexError:
                pass

        # BaşariOrani sütunundaki en büyük değeri al ve Sonuç sütununa ekle
        try:
            basari_orani_max = df["BaşariOrani"].max()
            basari_orani_max_id = df[df["BaşariOrani"] == basari_orani_max]["ID"].iloc[0]
            yeni_df.loc[len(yeni_df), "Sonuç"] = basari_orani_max_id
            df.drop(df[df["ID"] == basari_orani_max_id].index, inplace=True, errors="ignore")
        except IndexError:
            pass

        # İndirimOrani sütunundaki en büyük değeri al ve Sonuç sütununa ekle
        try:
            indirim_orani_max = df["İndirimOrani"].max()
            indirim_orani_max_id = df[df["İndirimOrani"] == indirim_orani_max]["ID"].iloc[0]
            yeni_df.loc[len(yeni_df), "Sonuç"] = indirim_orani_max_id
            df.drop(df[df["ID"] == indirim_orani_max_id].index, inplace=True, errors="ignore")
        except IndexError:
            pass

        # Görüntülenme sütunundaki ilk "Az Görüntülenme" değerini al ve Sonuç sütununa ekle
        try:
            az_goruntulenme_first = df[df["Görüntülenme"] == "Az Görüntülenme"]["ID"].iloc[0]
            yeni_df.loc[len(yeni_df), "Sonuç"] = az_goruntulenme_first
            df.drop(df[df["ID"] == az_goruntulenme_first].index, inplace=True, errors="ignore")
        except IndexError:
            pass

        # ID sütunundaki en büyük değeri al ve Sonuç sütununa ekle
        try:
            max_id = df["ID"].max()
            yeni_df.loc[len(yeni_df), "Sonuç"] = max_id
            df.drop(df[df["ID"] == max_id].index, inplace=True, errors="ignore")
        except KeyError:
            pass

    # Yeni Excel dosyasına yaz
    yeni_df.to_excel(yeni_birlesmis_veri, sheet_name=sheet_name, index=False)

# Yeni Excel dosyasını kaydet
yeni_birlesmis_veri.close()





# Excel dosyasını oku
df = pd.read_excel('yeni_birlesmis__veri.xlsx', sheet_name=None)

# "Sheet1" sayfasını sil
df.pop('Sheet1', None)

# Yeni bir Excel dosyası olarak kaydet
with pd.ExcelWriter('yeni_birlesmis__veri.xlsx', engine='openpyxl') as writer:
    for sheet_name, df_sheet in df.items():
        df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)





# "Birlestirilmis_Veriler_Sirali" Excel dosyasını oku
df_birlestirilmis_sirali = pd.read_excel('yeni_birlesmis__veri.xlsx', sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter('yeni_birlesmis__veri.xlsx', engine='openpyxl') as writer:
    # Her bir sayfa için işlemleri gerçekleştir
    for sheet_name, df_sheet_birlestirilmis_sirali in df_birlestirilmis_sirali.items():
        # SayfaIsmi sütunundaki tüm verileri "466" ile doldur
        df_sheet_birlestirilmis_sirali['Kategori ID'] = 502

        # Veriyi güncellenmiş haliyle aynı sayfaya kaydet
        df_sheet_birlestirilmis_sirali.to_excel(writer, sheet_name=sheet_name, index=False)



# "Birlestirilmis_Veriler_Sirali_Yeni" Excel dosyasını oku
df_birlestirilmis_sirali_yeni = pd.read_excel('yeni_birlesmis__veri.xlsx', sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter('yeni_birlesmis__veri.xlsx', engine='openpyxl') as writer:
    # Her bir sayfa için işlemleri gerçekleştir
    for sheet_name, df_sheet_birlestirilmis_sirali_yeni in df_birlestirilmis_sirali_yeni.items():
        # Her satıra sırasıyla -50, -49, -48, ... şeklinde sayıları ekle
        df_sheet_birlestirilmis_sirali_yeni['Numara'] = range(-len(df_sheet_birlestirilmis_sirali_yeni), 0)

        # Veriyi güncellenmiş haliyle aynı sayfaya kaydet
        df_sheet_birlestirilmis_sirali_yeni.to_excel(writer, sheet_name=sheet_name, index=False)







# Excel dosyasının adı ve konumu
excel_dosyasi = "yeni_birlesmis__veri.xlsx"

# Excel dosyasını yükle
birlesmis_veri = pd.ExcelFile(excel_dosyasi)

# Tüm sayfaların verilerini birleştirmek için boş bir DataFrame oluştur
birlesmis_df = pd.DataFrame()

# Her bir sayfa için işlem yap
for sayfa in birlesmis_veri.sheet_names:
    # Sayfa verisini oku
    veri = birlesmis_veri.parse(sayfa)
    
    # Her sayfanın verisini birleştir
    birlesmis_df = pd.concat([birlesmis_df, veri])

# Yeni bir dosya adı
yeni_dosya_adı = "Kategori Sıralama.xlsx"

# Birleştirilmiş veriyi yeni bir Excel dosyası olarak kaydet
birlesmis_df.to_excel(yeni_dosya_adı, index=False)





# Excel dosyasının adı ve konumu
excel_dosyasi = "Kategori Sıralama.xlsx"

# Excel dosyasını yükle ve DataFrame'e dönüştür
birlesmis_veri = pd.read_excel(excel_dosyasi)

# "Sonuç" sütununun adını "ID" olarak değiştir
birlesmis_veri.rename(columns={"Sonuç": "ID"}, inplace=True)

# Yeni bir dosya adı
yeni_dosya_adı = "Kategori Sıralama.xlsx"

# Güncellenmiş veriyi yeni bir Excel dosyası olarak kaydet
birlesmis_veri.to_excel(yeni_dosya_adı, index=False)




gc.collect()


# Silmek istediğiniz dosyaların listesi
dosya_listesi = ['yeni_birlesmis__veri.xlsx', 'birlesmis__veri.xlsx']

# Dosyaları silme işlemi
for dosya in dosya_listesi:
    try:
        os.remove(dosya)
    except FileNotFoundError:
        print(f"'{dosya}' dosyası bulunamadı.")
    except Exception as e:
        print(f"'{dosya}' dosyasını silerken bir hata oluştu: {str(e)}")












# Global tken değişkeni
_auth_token = None

# Token alma fonksiyonu
def get_auth_token():
    global _auth_token
    if _auth_token is None:  
        login_url = "https://siparis.haydigiy.com/api/customer/login"
        login_payload = {
            "apiKey": "MypGcaEInEOTzuYQydgDHQ",
            "secretKey": "jRqliBLDPke76YhL_WL5qg",
            "emailOrPhone": "mustafa_kod@haydigiy.com",
            "password": "123456"
        }
        login_headers = {
            "Content-Type": "application/json"
        }

        response = requests.post(login_url, json=login_payload, headers=login_headers)
        if response.status_code == 200:
            _auth_token = response.json().get("data", {}).get("token")
            if not _auth_token:
                raise Exception("TOKEN ALINAMADI")
        else:
            raise Exception(f"GİRİŞ BAŞARISIZ: {response.text}")
    return _auth_token

# Token alma işlemi
token = get_auth_token()
df = pd.read_excel("Kategori Sıralama.xlsx")
conn = http.client.HTTPSConnection("siparis.haydigiy.com")


for index, row in tqdm(df.iterrows(), total=df.shape[0], desc="Tesettür Kategorisi Sıralanıyor 1"):

    category_id = row['Kategori ID']
    display_order = row['Numara']
    product_id = row['ID']


    product_id = str(row['ID']).replace(".0", "")

    payload = json.dumps({
        "CategoryId": int(category_id),  
        "IsFeaturedProduct": False, 
        "DisplayOrder": int(display_order)
    })

    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {token}',
        'Cookie': '.Application.Customer=64684894-1b54-488d-bd59-76b94842df65'
    }


    conn.request("PUT", f"/adminapi/product/product-categories?productId={product_id}", payload, headers)
    res = conn.getresponse()
    data = res.read()



conn.close()


gc.collect()

# Silmek istediğiniz dosyaların listesi
dosya_listesi = ['Kategori Sıralama.xlsx']

# Dosyaları silme işlemi
for dosya in dosya_listesi:
    try:
        os.remove(dosya)
    
    except FileNotFoundError:
        print(f"'{dosya}' dosyası bulunamadı.")
    except Exception as e:
        print(f"'{dosya}' dosyasını silerken bir hata oluştu: {str(e)}")

#endregion

#region Tesettür Kategorisi Kırık Bedenler Sıralama

def get_excel_data(url):
    response = requests.get(url)

    if response.status_code == 200:
        # Excel dosyasını oku
        df = pd.read_excel(BytesIO(response.content))
        return df
    else:
        
        return None

# İlk linkten veriyi al
url1 = "https://task.haydigiy.com/FaprikaXls/5DO7BZ/1/"
data1 = get_excel_data(url1)

# İkinci linkten veriyi al
url2 = "https://task.haydigiy.com/FaprikaXls/5DO7BZ/2/"
data2 = get_excel_data(url2)

# İki veriyi birleştir
if data1 is not None and data2 is not None:
    merged_data = pd.concat([data1, data2], ignore_index=True)

    # Gereksiz sütunları sil
    columns_to_keep = ["StokAdedi", "UrunAdi", "AlisFiyati", "SatisFiyati", "Kategori", "MorhipoKodu", "HepsiBuradaKodu"]
    merged_data = merged_data[columns_to_keep]

    # Birleştirilmiş veriyi Excel dosyasına kaydet
    merged_data.to_excel("birlesmis__veri.xlsx", index=False)

    # Birleştirilmiş veriyi oku
    final_data = pd.read_excel("birlesmis__veri.xlsx")

else:
    pass


# Veriyi Okuma
df = pd.read_excel('birlesmis__veri.xlsx')

# Alış Fiyatına Göre İşlemler ve Kategori Kontrolü
def calculate_list_price(row):
    alis_fiyati = row['AlisFiyati']
    kategori = row['Kategori']

    if 0 <= alis_fiyati <= 24.99:
        result = alis_fiyati + 10
    elif 25 <= alis_fiyati <= 39.99:
        result = alis_fiyati + 13
    elif 40 <= alis_fiyati <= 59.99:
        result = alis_fiyati + 17
    elif 60 <= alis_fiyati <= 200.99:
        result = alis_fiyati * 1.30
    elif alis_fiyati >= 201:
        result = alis_fiyati * 1.25
    else:
        result = alis_fiyati

    # KDV
    if isinstance(kategori, str) and any(category in kategori for category in ["Parfüm", "Gözlük", "Saat"]):
        result *= 1.20
    else:
        result *= 1.10

    return result

# Yeni Sütun Oluşturma
df['ListeFiyati'] = df.apply(calculate_list_price, axis=1)

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('birlesmis__veri.xlsx', index=False)











# Veriyi Okuma
df = pd.read_excel('birlesmis__veri.xlsx')

# "StokAdedi" Sütununda 0'dan Büyük Olan Değerlerin Adedi
df['StokAdedi_GT_0'] = df['StokAdedi'].apply(lambda x: 1 if x > 0 else 0)
stok_adedi_gt_0_adet = df.groupby('UrunAdi')['StokAdedi_GT_0'].sum().reset_index()

# "UrunAdi" Sütunundaki Toplam Yenilenme Adedi
toplam_yenilenme_adedi = df.groupby('UrunAdi').size().reset_index(name='ToplamYenilenmeAdedi')

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0'], axis=1, errors='ignore')


# Oranı Hesapla ve Yeni Sütunu Ekle
df = pd.merge(df, stok_adedi_gt_0_adet, on='UrunAdi', how='left')
df = pd.merge(df, toplam_yenilenme_adedi, on='UrunAdi', how='left')
df['Beden Durumu'] = df['StokAdedi_GT_0'] / df['ToplamYenilenmeAdedi']

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0', 'ToplamYenilenmeAdedi'], axis=1, errors='ignore')

# Oranı 100 ile Çarpma
df['Beden Durumu'] *= 100

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('birlesmis__veri.xlsx', index=False)











# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# "ListeFiyati" ve "SatisFiyati" sütunlarındaki verilerden işlem yap
birlesmis_veri["İndirimOrani"] = (birlesmis_veri["ListeFiyati"] - birlesmis_veri["SatisFiyati"]) * 100 / birlesmis_veri["ListeFiyati"]

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)








# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# İstenmeyen sütunları sil
silinecek_sutunlar = ["StokAdedi", "AlisFiyati", "SatisFiyati", "ListeFiyati"]
birlesmis_veri = birlesmis_veri.drop(columns=silinecek_sutunlar, errors='ignore')

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)







# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# Tüm tablodaki benzersiz değerleri teke düşür
birlesmis_veri = birlesmis_veri.drop_duplicates()

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)






# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# "Beden Durumu" sütunundaki değeri 50'den küçük olan satırları filtrele
birlesmis_veri = birlesmis_veri[birlesmis_veri["Beden Durumu"] < 50]

# İstenmeyen sütunları sil
silinecek_sutunlar = ["Beden Durumu"]
birlesmis_veri = birlesmis_veri.drop(columns=silinecek_sutunlar, errors='ignore')

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)







# XML'den Ürün Bilgilerini Çekme ve Temizleme
xml_url = "https://task.haydigiy.com/FaprikaXml/1YJGSU/1/"
response = requests.get(xml_url)
xml_data = response.text
soup = BeautifulSoup(xml_data, 'xml')

product_data = []
for item in soup.find_all('item'):
    title = item.find('title').text.replace(' - Haydigiy', '')
    product_id = item.find('g:id').text if item.find('g:id') else None
    product_data.append({'UrunAdi': title, 'ID': product_id})

df_xml = pd.DataFrame(product_data)


# Excel ile Birleştirme
df_calisma_alani = pd.read_excel('birlesmis__veri.xlsx')
df_merged = pd.merge(df_calisma_alani, df_xml, how='left', left_on='UrunAdi', right_on='UrunAdi')

# Sonuçları Mevcut Excel Dosyasının Üzerine Kaydetme
df_merged.to_excel('birlesmis__veri.xlsx', index=False)






# birlesmis_veri Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx', sheet_name='Sheet1')

# "ID" sütunundaki boş olan hücreleri içeren satırları filtrele
df_birlesmis_veri = df_birlesmis_veri.dropna(subset=['ID'])

# Sonuçları güncellenmiş haliyle aynı Excel dosyasına kaydet
df_birlesmis_veri.to_excel('birlesmis__veri.xlsx', sheet_name='Sheet1', index=False)





# Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx')

# 'UrunAdi' sütununu sil
df_birlesmis_veri = df_birlesmis_veri.drop(columns=['UrunAdi'], errors='ignore')

# 'ID' sütununu en başa al
df_birlesmis_veri = df_birlesmis_veri[['ID'] + [col for col in df_birlesmis_veri.columns if col != 'ID']]

# Veriyi Excel dosyasına kaydet (üzerine yaz)
df_birlesmis_veri.to_excel('birlesmis__veri.xlsx', index=False)









# Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx')

# 'Kategori' sütununda "Tekrar Stokta" içeren hücreleri filtrele
df_birlesmis_veri['Tekrar Stokta'] = df_birlesmis_veri['Kategori'].apply(lambda x: 'Tekrar Stokta' if 'Tekrar Stokta' in str(x) else None)

# Veriyi Excel dosyasına kaydet (üzerine yaz)
df_birlesmis_veri.to_excel('birlesmis__veri.xlsx', index=False)








# Excel dosyasını oku
df = pd.read_excel('birlesmis__veri.xlsx')

# 'MorhipoKodu' sütunundaki değerleri sayıya dönüştür ve boş olan hücrelere "0" yaz
df['MorhipoKodu'] = pd.to_numeric(df['MorhipoKodu'], errors='coerce').fillna(0).astype(int)

# 'MorhipoKodu' sütunundaki değerleri sayıya dönüştür ve boş olan hücrelere "0" yaz
df['HepsiBuradaKodu'] = pd.to_numeric(df['HepsiBuradaKodu'], errors='coerce').fillna(0).astype(int)

# Güncellenmiş DataFrame'i aynı Excel dosyasının üzerine yaz
df.to_excel('birlesmis__veri.xlsx', index=False)











# Excel dosyasını oku
df = pd.read_excel('birlesmis__veri.xlsx')

# 'BaşariOrani' sütununu oluştur ve işlemi gerçekleştir
df['BaşariOrani'] = df['MorhipoKodu'] / df['HepsiBuradaKodu'].replace(0, 1)

# 'HepsiBuradaKodu' sütunundaki 0 değerlerini 'BaşariOrani' sütununa 0 olarak yaz
df.loc[df['HepsiBuradaKodu'] == 0, 'BaşariOrani'] = 0

# 'İndirimOrani' sütunundaki değerlerde 1'den küçük olanları "0" ile değiştir
df.loc[df['İndirimOrani'] < 1, 'İndirimOrani'] = 0

# Güncellenmiş DataFrame'i aynı Excel dosyasının üzerine yaz
df.to_excel('birlesmis__veri.xlsx', index=False)







# Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx')

# İstenilen kategorileri içeren satırları filtrele
istenen_kategoriler = [
    'TESETTÜR'
]

for kategori in istenen_kategoriler:
    df_kategori = df_birlesmis_veri[df_birlesmis_veri['Kategori'].str.contains(kategori, case=False, na=False)]
    
    # Veriyi yeni bir sayfaya kaydet (excel dosyasının içine eklenir)
    with pd.ExcelWriter('birlesmis__veri.xlsx', engine='openpyxl', mode='a') as writer:
        df_kategori.to_excel(writer, sheet_name=f'{kategori}', index=False)








# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx", sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter("birlesmis__veri.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    # Her sayfa için işlemi yap
    for sheet_name, df in birlesmis_veri.items():
        # Yeni sütunu hesapla
        df["Görüntülenme"] = (100 *  df["HepsiBuradaKodu"] / df["HepsiBuradaKodu"].sum())

        # Sonucu Excel dosyasına yaz
        df.to_excel(writer, sheet_name=sheet_name, index=False)






# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx", sheet_name=None)

# Her sayfa için işlemi yap
for sheet_name, df in birlesmis_veri.items():
    # "Görüntülenme" sütununu sayısal değerlere dönüştür
    df["Görüntülenme"] = pd.to_numeric(df["Görüntülenme"], errors='coerce')
    
    # Yeni sütunu hesapla
    ortalama = 0.7 * df["Görüntülenme"].mean()
    
    # Değerleri güncelle
    df["Görüntülenme"] = np.where(df["Görüntülenme"] < ortalama, "Az Görüntülenme", "")


    df.drop(columns=["HepsiBuradaKodu"], inplace=True)

    # Excel dosyasına yaz
    with pd.ExcelWriter("birlesmis__veri.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)







# Yeni bir Excel dosyası oluştur
yeni_birlesmis_veri = pd.ExcelWriter("yeni_birlesmis__veri.xlsx")

# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx", sheet_name=None)

# Her sayfa için işlemi yap
for sheet_name, df in birlesmis_veri.items():
    # Eğer veri yoksa işlem yapma
    if df.empty:
        continue

    # Yeni bir veri çerçevesi oluştur
    yeni_df = pd.DataFrame()

    # MorhipoKodu sütunundaki en büyük ilk 6 değeri al ve Sonuç sütununa ekle
    morhipo_kodu_top_6 = df.nlargest(6, "MorhipoKodu")
    for index, row in morhipo_kodu_top_6.iterrows():
        yeni_df.loc[len(yeni_df), "Sonuç"] = row["ID"]
    df.drop(morhipo_kodu_top_6.index, inplace=True, errors="ignore")

    while not df.empty:
        # Tekrar Stokta sütununda "Tekrar Stokta" verisi var mı kontrol et
        if "Tekrar Stokta" in df.columns and "Tekrar Stokta" in df["Tekrar Stokta"].unique():
            # Tekrar Stokta sütunundaki ilk "Tekrar Stokta" değerini al ve Sonuç sütununa ekle
            try:
                tekrar_stokta_first = df[df["Tekrar Stokta"] == "Tekrar Stokta"]["ID"].iloc[0]
                yeni_df.loc[len(yeni_df), "Sonuç"] = tekrar_stokta_first
                df.drop(df[df["ID"] == tekrar_stokta_first].index, inplace=True, errors="ignore")
            except IndexError:
                pass

        # BaşariOrani sütunundaki en büyük değeri al ve Sonuç sütununa ekle
        try:
            basari_orani_max = df["BaşariOrani"].max()
            basari_orani_max_id = df[df["BaşariOrani"] == basari_orani_max]["ID"].iloc[0]
            yeni_df.loc[len(yeni_df), "Sonuç"] = basari_orani_max_id
            df.drop(df[df["ID"] == basari_orani_max_id].index, inplace=True, errors="ignore")
        except IndexError:
            pass

        # İndirimOrani sütunundaki en büyük değeri al ve Sonuç sütununa ekle
        try:
            indirim_orani_max = df["İndirimOrani"].max()
            indirim_orani_max_id = df[df["İndirimOrani"] == indirim_orani_max]["ID"].iloc[0]
            yeni_df.loc[len(yeni_df), "Sonuç"] = indirim_orani_max_id
            df.drop(df[df["ID"] == indirim_orani_max_id].index, inplace=True, errors="ignore")
        except IndexError:
            pass

        # Görüntülenme sütunundaki ilk "Az Görüntülenme" değerini al ve Sonuç sütununa ekle
        try:
            az_goruntulenme_first = df[df["Görüntülenme"] == "Az Görüntülenme"]["ID"].iloc[0]
            yeni_df.loc[len(yeni_df), "Sonuç"] = az_goruntulenme_first
            df.drop(df[df["ID"] == az_goruntulenme_first].index, inplace=True, errors="ignore")
        except IndexError:
            pass

        # ID sütunundaki en büyük değeri al ve Sonuç sütununa ekle
        try:
            max_id = df["ID"].max()
            yeni_df.loc[len(yeni_df), "Sonuç"] = max_id
            df.drop(df[df["ID"] == max_id].index, inplace=True, errors="ignore")
        except KeyError:
            pass

    # Yeni Excel dosyasına yaz
    yeni_df.to_excel(yeni_birlesmis_veri, sheet_name=sheet_name, index=False)

# Yeni Excel dosyasını kaydet
yeni_birlesmis_veri.close()





# Excel dosyasını oku
df = pd.read_excel('yeni_birlesmis__veri.xlsx', sheet_name=None)

# "Sheet1" sayfasını sil
df.pop('Sheet1', None)

# Yeni bir Excel dosyası olarak kaydet
with pd.ExcelWriter('yeni_birlesmis__veri.xlsx', engine='openpyxl') as writer:
    for sheet_name, df_sheet in df.items():
        df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)





# "Birlestirilmis_Veriler_Sirali" Excel dosyasını oku
df_birlestirilmis_sirali = pd.read_excel('yeni_birlesmis__veri.xlsx', sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter('yeni_birlesmis__veri.xlsx', engine='openpyxl') as writer:
    # Her bir sayfa için işlemleri gerçekleştir
    for sheet_name, df_sheet_birlestirilmis_sirali in df_birlestirilmis_sirali.items():
        # SayfaIsmi sütunundaki tüm verileri "466" ile doldur
        df_sheet_birlestirilmis_sirali['Kategori ID'] = 502

        # Veriyi güncellenmiş haliyle aynı sayfaya kaydet
        df_sheet_birlestirilmis_sirali.to_excel(writer, sheet_name=sheet_name, index=False)




# "Birlestirilmis_Veriler_Sirali_Yeni" Excel dosyasını oku
df_birlestirilmis_sirali_yeni = pd.read_excel('yeni_birlesmis__veri.xlsx', sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter('yeni_birlesmis__veri.xlsx', engine='openpyxl') as writer:
    # Her bir sayfa için işlemleri gerçekleştir
    for sheet_name, df_sheet_birlestirilmis_sirali_yeni in df_birlestirilmis_sirali_yeni.items():
        # Her satıra sırasıyla -50, -49, -48, ... şeklinde sayıları ekle
        df_sheet_birlestirilmis_sirali_yeni['Numara'] = range(-len(df_sheet_birlestirilmis_sirali_yeni), 0)

        # "Numara" sütunundaki tüm verileri 0 ile değiştir
        df_sheet_birlestirilmis_sirali_yeni['Numara'] = 0

        # Veriyi güncellenmiş haliyle aynı sayfaya kaydet
        df_sheet_birlestirilmis_sirali_yeni.to_excel(writer, sheet_name=sheet_name, index=False)












# Excel dosyasının adı ve konumu
excel_dosyasi = "yeni_birlesmis__veri.xlsx"

# Excel dosyasını yükle
birlesmis_veri = pd.ExcelFile(excel_dosyasi)

# Tüm sayfaların verilerini birleştirmek için boş bir DataFrame oluştur
birlesmis_df = pd.DataFrame()

# Her bir sayfa için işlem yap
for sayfa in birlesmis_veri.sheet_names:
    # Sayfa verisini oku
    veri = birlesmis_veri.parse(sayfa)
    
    # Her sayfanın verisini birleştir
    birlesmis_df = pd.concat([birlesmis_df, veri])

# Yeni bir dosya adı
yeni_dosya_adı = "Kategori Sıralama.xlsx"

# Birleştirilmiş veriyi yeni bir Excel dosyası olarak kaydet
birlesmis_df.to_excel(yeni_dosya_adı, index=False)





# Excel dosyasının adı ve konumu
excel_dosyasi = "Kategori Sıralama.xlsx"

# Excel dosyasını yükle ve DataFrame'e dönüştür
birlesmis_veri = pd.read_excel(excel_dosyasi)

# "Sonuç" sütununun adını "ID" olarak değiştir
birlesmis_veri.rename(columns={"Sonuç": "ID"}, inplace=True)

# Yeni bir dosya adı
yeni_dosya_adı = "Kategori Sıralama.xlsx"

# Güncellenmiş veriyi yeni bir Excel dosyası olarak kaydet
birlesmis_veri.to_excel(yeni_dosya_adı, index=False)





gc.collect()


# Silmek istediğiniz dosyaların listesi
dosya_listesi = ['yeni_birlesmis__veri.xlsx', 'birlesmis__veri.xlsx']

# Dosyaları silme işlemi
for dosya in dosya_listesi:
    try:
        os.remove(dosya)
    except FileNotFoundError:
        print(f"'{dosya}' dosyası bulunamadı.")
    except Exception as e:
        print(f"'{dosya}' dosyasını silerken bir hata oluştu: {str(e)}")










# Global tken değişkeni
_auth_token = None

# Token alma fonksiyonu
def get_auth_token():
    global _auth_token
    if _auth_token is None:  
        login_url = "https://siparis.haydigiy.com/api/customer/login"
        login_payload = {
            "apiKey": "MypGcaEInEOTzuYQydgDHQ",
            "secretKey": "jRqliBLDPke76YhL_WL5qg",
            "emailOrPhone": "mustafa_kod@haydigiy.com",
            "password": "123456"
        }
        login_headers = {
            "Content-Type": "application/json"
        }

        response = requests.post(login_url, json=login_payload, headers=login_headers)
        if response.status_code == 200:
            _auth_token = response.json().get("data", {}).get("token")
            if not _auth_token:
                raise Exception("TOKEN ALINAMADI")
        else:
            raise Exception(f"GİRİŞ BAŞARISIZ: {response.text}")
    return _auth_token

# Token alma işlemi
token = get_auth_token()
df = pd.read_excel("Kategori Sıralama.xlsx")
conn = http.client.HTTPSConnection("siparis.haydigiy.com")


for index, row in tqdm(df.iterrows(), total=df.shape[0], desc="Tesettür Kategorisi Sıralanıyor 2"):

    category_id = row['Kategori ID']
    display_order = row['Numara']
    product_id = row['ID']


    product_id = str(row['ID']).replace(".0", "")

    payload = json.dumps({
        "CategoryId": int(category_id),  
        "IsFeaturedProduct": False, 
        "DisplayOrder": int(display_order)
    })

    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {token}',
        'Cookie': '.Application.Customer=64684894-1b54-488d-bd59-76b94842df65'
    }


    conn.request("PUT", f"/adminapi/product/product-categories?productId={product_id}", payload, headers)
    res = conn.getresponse()
    data = res.read()

conn.close()




# Silmek istediğiniz dosyaların listesi
dosya_listesi = ['Kategori Sıralama.xlsx']

# Dosyaları silme işlemi
for dosya in dosya_listesi:
    try:
        os.remove(dosya)
    
    except FileNotFoundError:
        print(f"'{dosya}' dosyası bulunamadı.")
    except Exception as e:
        print(f"'{dosya}' dosyasını silerken bir hata oluştu: {str(e)}")

#endregion

#region Tesettür Kategorisi Diğer Ürünler Sıralama

def get_excel_data(url):
    response = requests.get(url)

    if response.status_code == 200:
        # Excel dosyasını oku
        df = pd.read_excel(BytesIO(response.content))
        return df
    else:
        
        return None

# İlk linkten veriyi al
url1 = "https://task.haydigiy.com/FaprikaXls/V4IYBO/1/"
data1 = get_excel_data(url1)

# İkinci linkten veriyi al
url2 = "https://task.haydigiy.com/FaprikaXls/V4IYBO/2/"
data2 = get_excel_data(url2)

# İki veriyi birleştir
if data1 is not None and data2 is not None:
    merged_data = pd.concat([data1, data2], ignore_index=True)

    # Gereksiz sütunları sil
    columns_to_keep = ["StokAdedi", "UrunAdi", "AlisFiyati", "SatisFiyati", "Kategori", "MorhipoKodu", "HepsiBuradaKodu"]
    merged_data = merged_data[columns_to_keep]

    # Birleştirilmiş veriyi Excel dosyasına kaydet
    merged_data.to_excel("birlesmis__veri.xlsx", index=False)

    # Birleştirilmiş veriyi oku
    final_data = pd.read_excel("birlesmis__veri.xlsx")

else:
    pass


# Veriyi Okuma
df = pd.read_excel('birlesmis__veri.xlsx')

# Alış Fiyatına Göre İşlemler ve Kategori Kontrolü
def calculate_list_price(row):
    alis_fiyati = row['AlisFiyati']
    kategori = row['Kategori']

    if 0 <= alis_fiyati <= 24.99:
        result = alis_fiyati + 10
    elif 25 <= alis_fiyati <= 39.99:
        result = alis_fiyati + 13
    elif 40 <= alis_fiyati <= 59.99:
        result = alis_fiyati + 17
    elif 60 <= alis_fiyati <= 200.99:
        result = alis_fiyati * 1.30
    elif alis_fiyati >= 201:
        result = alis_fiyati * 1.25
    else:
        result = alis_fiyati

    # KDV
    if isinstance(kategori, str) and any(category in kategori for category in ["Parfüm", "Gözlük", "Saat"]):
        result *= 1.20
    else:
        result *= 1.10

    return result


# Yeni Sütun Oluşturma
df['ListeFiyati'] = df.apply(calculate_list_price, axis=1)

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('birlesmis__veri.xlsx', index=False)











# Veriyi Okuma
df = pd.read_excel('birlesmis__veri.xlsx')

# "StokAdedi" Sütununda 0'dan Büyük Olan Değerlerin Adedi
df['StokAdedi_GT_0'] = df['StokAdedi'].apply(lambda x: 1 if x > 0 else 0)
stok_adedi_gt_0_adet = df.groupby('UrunAdi')['StokAdedi_GT_0'].sum().reset_index()

# "UrunAdi" Sütunundaki Toplam Yenilenme Adedi
toplam_yenilenme_adedi = df.groupby('UrunAdi').size().reset_index(name='ToplamYenilenmeAdedi')

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0'], axis=1, errors='ignore')


# Oranı Hesapla ve Yeni Sütunu Ekle
df = pd.merge(df, stok_adedi_gt_0_adet, on='UrunAdi', how='left')
df = pd.merge(df, toplam_yenilenme_adedi, on='UrunAdi', how='left')
df['Beden Durumu'] = df['StokAdedi_GT_0'] / df['ToplamYenilenmeAdedi']

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0', 'ToplamYenilenmeAdedi'], axis=1, errors='ignore')

# Oranı 100 ile Çarpma
df['Beden Durumu'] *= 100

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('birlesmis__veri.xlsx', index=False)











# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# "ListeFiyati" ve "SatisFiyati" sütunlarındaki verilerden işlem yap
birlesmis_veri["İndirimOrani"] = (birlesmis_veri["ListeFiyati"] - birlesmis_veri["SatisFiyati"]) * 100 / birlesmis_veri["ListeFiyati"]

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)








# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# İstenmeyen sütunları sil
silinecek_sutunlar = ["StokAdedi", "AlisFiyati", "SatisFiyati", "ListeFiyati"]
birlesmis_veri = birlesmis_veri.drop(columns=silinecek_sutunlar, errors='ignore')

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)







# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# Tüm tablodaki benzersiz değerleri teke düşür
birlesmis_veri = birlesmis_veri.drop_duplicates()

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)






# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx")

# "Beden Durumu" sütunundaki değeri 50'den küçük olan satırları filtrele
birlesmis_veri = birlesmis_veri[birlesmis_veri["Beden Durumu"] < 50]

# İstenmeyen sütunları sil
silinecek_sutunlar = ["Beden Durumu"]
birlesmis_veri = birlesmis_veri.drop(columns=silinecek_sutunlar, errors='ignore')

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("birlesmis__veri.xlsx", index=False)







# XML'den Ürün Bilgilerini Çekme ve Temizleme
xml_url = "https://task.haydigiy.com/FaprikaXml/9RN1ZY/1/"
response = requests.get(xml_url)
xml_data = response.text
soup = BeautifulSoup(xml_data, 'xml')

product_data = []
for item in soup.find_all('item'):
    title = item.find('title').text.replace(' - Haydigiy', '')
    product_id = item.find('g:id').text if item.find('g:id') else None
    product_data.append({'UrunAdi': title, 'ID': product_id})

df_xml = pd.DataFrame(product_data)


# Excel ile Birleştirme
df_calisma_alani = pd.read_excel('birlesmis__veri.xlsx')
df_merged = pd.merge(df_calisma_alani, df_xml, how='left', left_on='UrunAdi', right_on='UrunAdi')

# Sonuçları Mevcut Excel Dosyasının Üzerine Kaydetme
df_merged.to_excel('birlesmis__veri.xlsx', index=False)






# birlesmis_veri Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx', sheet_name='Sheet1')

# "ID" sütunundaki boş olan hücreleri içeren satırları filtrele
df_birlesmis_veri = df_birlesmis_veri.dropna(subset=['ID'])

# Sonuçları güncellenmiş haliyle aynı Excel dosyasına kaydet
df_birlesmis_veri.to_excel('birlesmis__veri.xlsx', sheet_name='Sheet1', index=False)





# Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx')

# 'UrunAdi' sütununu sil
df_birlesmis_veri = df_birlesmis_veri.drop(columns=['UrunAdi'], errors='ignore')

# 'ID' sütununu en başa al
df_birlesmis_veri = df_birlesmis_veri[['ID'] + [col for col in df_birlesmis_veri.columns if col != 'ID']]

# Veriyi Excel dosyasına kaydet (üzerine yaz)
df_birlesmis_veri.to_excel('birlesmis__veri.xlsx', index=False)









# Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx')

# 'Kategori' sütununda "Tekrar Stokta" içeren hücreleri filtrele
df_birlesmis_veri['Tekrar Stokta'] = df_birlesmis_veri['Kategori'].apply(lambda x: 'Tekrar Stokta' if 'Tekrar Stokta' in str(x) else None)

# Veriyi Excel dosyasına kaydet (üzerine yaz)
df_birlesmis_veri.to_excel('birlesmis__veri.xlsx', index=False)








# Excel dosyasını oku
df = pd.read_excel('birlesmis__veri.xlsx')

# 'MorhipoKodu' sütunundaki değerleri sayıya dönüştür ve boş olan hücrelere "0" yaz
df['MorhipoKodu'] = pd.to_numeric(df['MorhipoKodu'], errors='coerce').fillna(0).astype(int)

# 'MorhipoKodu' sütunundaki değerleri sayıya dönüştür ve boş olan hücrelere "0" yaz
df['HepsiBuradaKodu'] = pd.to_numeric(df['HepsiBuradaKodu'], errors='coerce').fillna(0).astype(int)

# Güncellenmiş DataFrame'i aynı Excel dosyasının üzerine yaz
df.to_excel('birlesmis__veri.xlsx', index=False)











# Excel dosyasını oku
df = pd.read_excel('birlesmis__veri.xlsx')

# 'BaşariOrani' sütununu oluştur ve işlemi gerçekleştir
df['BaşariOrani'] = df['MorhipoKodu'] / df['HepsiBuradaKodu'].replace(0, 1)

# 'HepsiBuradaKodu' sütunundaki 0 değerlerini 'BaşariOrani' sütununa 0 olarak yaz
df.loc[df['HepsiBuradaKodu'] == 0, 'BaşariOrani'] = 0

# 'İndirimOrani' sütunundaki değerlerde 1'den küçük olanları "0" ile değiştir
df.loc[df['İndirimOrani'] < 1, 'İndirimOrani'] = 0

# Güncellenmiş DataFrame'i aynı Excel dosyasının üzerine yaz
df.to_excel('birlesmis__veri.xlsx', index=False)







# Excel dosyasını oku
df_birlesmis_veri = pd.read_excel('birlesmis__veri.xlsx')

# İstenilen kategorileri içeren satırları filtrele
istenen_kategoriler = [
    'TESETTÜR'
]

for kategori in istenen_kategoriler:
    df_kategori = df_birlesmis_veri[df_birlesmis_veri['Kategori'].str.contains(kategori, case=False, na=False)]
    
    # Veriyi yeni bir sayfaya kaydet (excel dosyasının içine eklenir)
    with pd.ExcelWriter('birlesmis__veri.xlsx', engine='openpyxl', mode='a') as writer:
        df_kategori.to_excel(writer, sheet_name=f'{kategori}', index=False)








# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx", sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter("birlesmis__veri.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    # Her sayfa için işlemi yap
    for sheet_name, df in birlesmis_veri.items():
        # Yeni sütunu hesapla
        df["Görüntülenme"] = (100 *  df["HepsiBuradaKodu"] / df["HepsiBuradaKodu"].sum())

        # Sonucu Excel dosyasına yaz
        df.to_excel(writer, sheet_name=sheet_name, index=False)






# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx", sheet_name=None)

# Her sayfa için işlemi yap
for sheet_name, df in birlesmis_veri.items():
    # "Görüntülenme" sütununu sayısal değerlere dönüştür
    df["Görüntülenme"] = pd.to_numeric(df["Görüntülenme"], errors='coerce')
    
    # Yeni sütunu hesapla
    ortalama = 0.7 * df["Görüntülenme"].mean()
    
    # Değerleri güncelle
    df["Görüntülenme"] = np.where(df["Görüntülenme"] < ortalama, "Az Görüntülenme", "")


    df.drop(columns=["HepsiBuradaKodu"], inplace=True)

    # Excel dosyasına yaz
    with pd.ExcelWriter("birlesmis__veri.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)







# Yeni bir Excel dosyası oluştur
yeni_birlesmis_veri = pd.ExcelWriter("yeni_birlesmis__veri.xlsx")

# Excel dosyasını oku
birlesmis_veri = pd.read_excel("birlesmis__veri.xlsx", sheet_name=None)

# Her sayfa için işlemi yap
for sheet_name, df in birlesmis_veri.items():
    # Eğer veri yoksa işlem yapma
    if df.empty:
        continue

    # Yeni bir veri çerçevesi oluştur
    yeni_df = pd.DataFrame()

    # MorhipoKodu sütunundaki en büyük ilk 6 değeri al ve Sonuç sütununa ekle
    morhipo_kodu_top_6 = df.nlargest(6, "MorhipoKodu")
    for index, row in morhipo_kodu_top_6.iterrows():
        yeni_df.loc[len(yeni_df), "Sonuç"] = row["ID"]
    df.drop(morhipo_kodu_top_6.index, inplace=True, errors="ignore")

    while not df.empty:
        # Tekrar Stokta sütununda "Tekrar Stokta" verisi var mı kontrol et
        if "Tekrar Stokta" in df.columns and "Tekrar Stokta" in df["Tekrar Stokta"].unique():
            # Tekrar Stokta sütunundaki ilk "Tekrar Stokta" değerini al ve Sonuç sütununa ekle
            try:
                tekrar_stokta_first = df[df["Tekrar Stokta"] == "Tekrar Stokta"]["ID"].iloc[0]
                yeni_df.loc[len(yeni_df), "Sonuç"] = tekrar_stokta_first
                df.drop(df[df["ID"] == tekrar_stokta_first].index, inplace=True, errors="ignore")
            except IndexError:
                pass

        # BaşariOrani sütunundaki en büyük değeri al ve Sonuç sütununa ekle
        try:
            basari_orani_max = df["BaşariOrani"].max()
            basari_orani_max_id = df[df["BaşariOrani"] == basari_orani_max]["ID"].iloc[0]
            yeni_df.loc[len(yeni_df), "Sonuç"] = basari_orani_max_id
            df.drop(df[df["ID"] == basari_orani_max_id].index, inplace=True, errors="ignore")
        except IndexError:
            pass

        # İndirimOrani sütunundaki en büyük değeri al ve Sonuç sütununa ekle
        try:
            indirim_orani_max = df["İndirimOrani"].max()
            indirim_orani_max_id = df[df["İndirimOrani"] == indirim_orani_max]["ID"].iloc[0]
            yeni_df.loc[len(yeni_df), "Sonuç"] = indirim_orani_max_id
            df.drop(df[df["ID"] == indirim_orani_max_id].index, inplace=True, errors="ignore")
        except IndexError:
            pass

        # Görüntülenme sütunundaki ilk "Az Görüntülenme" değerini al ve Sonuç sütununa ekle
        try:
            az_goruntulenme_first = df[df["Görüntülenme"] == "Az Görüntülenme"]["ID"].iloc[0]
            yeni_df.loc[len(yeni_df), "Sonuç"] = az_goruntulenme_first
            df.drop(df[df["ID"] == az_goruntulenme_first].index, inplace=True, errors="ignore")
        except IndexError:
            pass

        # ID sütunundaki en büyük değeri al ve Sonuç sütununa ekle
        try:
            max_id = df["ID"].max()
            yeni_df.loc[len(yeni_df), "Sonuç"] = max_id
            df.drop(df[df["ID"] == max_id].index, inplace=True, errors="ignore")
        except KeyError:
            pass

    # Yeni Excel dosyasına yaz
    yeni_df.to_excel(yeni_birlesmis_veri, sheet_name=sheet_name, index=False)

# Yeni Excel dosyasını kaydet
yeni_birlesmis_veri.close()





# Excel dosyasını oku
df = pd.read_excel('yeni_birlesmis__veri.xlsx', sheet_name=None)

# "Sheet1" sayfasını sil
df.pop('Sheet1', None)

# Yeni bir Excel dosyası olarak kaydet
with pd.ExcelWriter('yeni_birlesmis__veri.xlsx', engine='openpyxl') as writer:
    for sheet_name, df_sheet in df.items():
        df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)




# "Birlestirilmis_Veriler_Sirali" Excel dosyasını oku
df_birlestirilmis_sirali = pd.read_excel('yeni_birlesmis__veri.xlsx', sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter('yeni_birlesmis__veri.xlsx', engine='openpyxl') as writer:
    # Her bir sayfa için işlemleri gerçekleştir
    for sheet_name, df_sheet_birlestirilmis_sirali in df_birlestirilmis_sirali.items():
        # SayfaIsmi sütunundaki tüm verileri "466" ile doldur
        df_sheet_birlestirilmis_sirali['Kategori ID'] = 502

        # Veriyi güncellenmiş haliyle aynı sayfaya kaydet
        df_sheet_birlestirilmis_sirali.to_excel(writer, sheet_name=sheet_name, index=False)




# "Birlestirilmis_Veriler_Sirali_Yeni" Excel dosyasını oku
df_birlestirilmis_sirali_yeni = pd.read_excel('yeni_birlesmis__veri.xlsx', sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter('yeni_birlesmis__veri.xlsx', engine='openpyxl') as writer:
    # Her bir sayfa için işlemleri gerçekleştir
    for sheet_name, df_sheet_birlestirilmis_sirali_yeni in df_birlestirilmis_sirali_yeni.items():
        # Her satıra sırasıyla -50, -49, -48, ... şeklinde sayıları ekle
        df_sheet_birlestirilmis_sirali_yeni['Numara'] = range(-len(df_sheet_birlestirilmis_sirali_yeni), 0)

        # "Numara" sütunundaki tüm verileri 0 ile değiştir
        df_sheet_birlestirilmis_sirali_yeni['Numara'] = 1000

        # Veriyi güncellenmiş haliyle aynı sayfaya kaydet
        df_sheet_birlestirilmis_sirali_yeni.to_excel(writer, sheet_name=sheet_name, index=False)












# Excel dosyasının adı ve konumu
excel_dosyasi = "yeni_birlesmis__veri.xlsx"

# Excel dosyasını yükle
birlesmis_veri = pd.ExcelFile(excel_dosyasi)

# Tüm sayfaların verilerini birleştirmek için boş bir DataFrame oluştur
birlesmis_df = pd.DataFrame()

# Her bir sayfa için işlem yap
for sayfa in birlesmis_veri.sheet_names:
    # Sayfa verisini oku
    veri = birlesmis_veri.parse(sayfa)
    
    # Her sayfanın verisini birleştir
    birlesmis_df = pd.concat([birlesmis_df, veri])

# Yeni bir dosya adı
yeni_dosya_adı = "Kategori Sıralama.xlsx"

# Birleştirilmiş veriyi yeni bir Excel dosyası olarak kaydet
birlesmis_df.to_excel(yeni_dosya_adı, index=False)





# Excel dosyasının adı ve konumu
excel_dosyasi = "Kategori Sıralama.xlsx"

# Excel dosyasını yükle ve DataFrame'e dönüştür
birlesmis_veri = pd.read_excel(excel_dosyasi)

# "Sonuç" sütununun adını "ID" olarak değiştir
birlesmis_veri.rename(columns={"Sonuç": "ID"}, inplace=True)

# Yeni bir dosya adı
yeni_dosya_adı = "Kategori Sıralama.xlsx"

# Güncellenmiş veriyi yeni bir Excel dosyası olarak kaydet
birlesmis_veri.to_excel(yeni_dosya_adı, index=False)






# Excel dosyasının adı ve konumu
excel_dosyasi = "Kategori Sıralama.xlsx"

# Excel dosyasını yükle
birlesmis_veri = pd.read_excel(excel_dosyasi)

# ID sütunundaki her verinin başına belirli bir metni ekleyerek yeni bir sütun oluştur
birlesmis_veri["ID"] = "https://task.haydigiy.com/admin/product/edit/" + birlesmis_veri["ID"].astype(str)

# Yeni bir dosya adı
yeni_dosya_adı = "Kategori Sıralama.xlsx"

# Güncellenmiş veriyi yeni bir Excel dosyası olarak kaydet
birlesmis_veri.to_excel(yeni_dosya_adı, index=False)






gc.collect()

# Silmek istediğiniz dosyaların listesi
dosya_listesi = ['yeni_birlesmis__veri.xlsx', 'birlesmis__veri.xlsx']

# Dosyaları silme işlemi
for dosya in dosya_listesi:
    try:
        os.remove(dosya)
    except FileNotFoundError:
        print(f"'{dosya}' dosyası bulunamadı.")
    except Exception as e:
        print(f"'{dosya}' dosyasını silerken bir hata oluştu: {str(e)}")











# Global tken değişkeni
_auth_token = None

# Token alma fonksiyonu
def get_auth_token():
    global _auth_token
    if _auth_token is None:  
        login_url = "https://siparis.haydigiy.com/api/customer/login"
        login_payload = {
            "apiKey": "MypGcaEInEOTzuYQydgDHQ",
            "secretKey": "jRqliBLDPke76YhL_WL5qg",
            "emailOrPhone": "mustafa_kod@haydigiy.com",
            "password": "123456"
        }
        login_headers = {
            "Content-Type": "application/json"
        }

        response = requests.post(login_url, json=login_payload, headers=login_headers)
        if response.status_code == 200:
            _auth_token = response.json().get("data", {}).get("token")
            if not _auth_token:
                raise Exception("TOKEN ALINAMADI")
        else:
            raise Exception(f"GİRİŞ BAŞARISIZ: {response.text}")
    return _auth_token

# Token alma işlemi
token = get_auth_token()
df = pd.read_excel("Kategori Sıralama.xlsx")
conn = http.client.HTTPSConnection("siparis.haydigiy.com")


for index, row in tqdm(df.iterrows(), total=df.shape[0], desc="Tesettür Kategorisi Sıralanıyor 3"):

    category_id = row['Kategori ID']
    display_order = row['Numara']
    product_id = row['ID']


    product_id = str(row['ID']).replace(".0", "")

    payload = json.dumps({
        "CategoryId": int(category_id),  
        "IsFeaturedProduct": False, 
        "DisplayOrder": int(display_order)
    })

    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {token}',
        'Cookie': '.Application.Customer=64684894-1b54-488d-bd59-76b94842df65'
    }


    conn.request("PUT", f"/adminapi/product/product-categories?productId={product_id}", payload, headers)
    res = conn.getresponse()
    data = res.read()

conn.close()



gc.collect()

# Silmek istediğiniz dosyaların listesi
dosya_listesi = ['Kategori Sıralama.xlsx']

# Dosyaları silme işlemi
for dosya in dosya_listesi:
    try:
        os.remove(dosya)
    
    except FileNotFoundError:
        print(f"'{dosya}' dosyası bulunamadı.")
    except Exception as e:
        print(f"'{dosya}' dosyasını silerken bir hata oluştu: {str(e)}")

#endregion




#region Ürün Listesi İndirme ve Gereksiz Sütunları Silme

def get_excel_data(url):
    response = requests.get(url)

    if response.status_code == 200:
        # Excel dosyasını oku
        df = pd.read_excel(BytesIO(response.content))
        return df
    else:
        return None

url1 = "https://task.haydigiy.com/FaprikaXls/I5UH6E/1/"
data1 = get_excel_data(url1)
url2 = "https://task.haydigiy.com/FaprikaXls/I5UH6E/2/"
data2 = get_excel_data(url2)

# İki veriyi birleştir
if data1 is not None and data2 is not None:
    merged_data = pd.concat([data1, data2], ignore_index=True)

    # İlk olarak Excel dosyasını kaydedelim
    original_file = "Stabil Ürün Listesi.xlsx"
    merged_data.to_excel(original_file, index=False)

    # Dosyanın bir kopyasını oluşturalım
    copy_file = "Öne Çıkanlar.xlsx"
    shutil.copy(original_file, copy_file)

    # Dosyanın bir kopyasını oluşturalım
    copy_file = "Öne Çıkanlar Yükleme.xlsx"
    shutil.copy(original_file, copy_file)

    # Stabil Ürün Listesi dosyasındaki gereksiz sütunları silelim
    columns_to_keep = ["StokAdedi", "UrunAdi", "AlisFiyati", "SatisFiyati", "Kategori", "MorhipoKodu", "HepsiBuradaKodu", "TrendyolKodu", "AramaTerimleri"]
    merged_data = merged_data[columns_to_keep]

    # Gereksiz sütunları sildikten sonra dosyayı yeniden kaydet
    merged_data.to_excel(original_file, index=False)

#endregion

#region Öne Çıkanlar Excel'inde Gereksiz Sütunları Silme

# Excel dosyasını okumak
df = pd.read_excel("Öne Çıkanlar.xlsx")

# Tutulacak sütunlar
columns_to_keep = ["UrunAdi", "AlisFiyati", "StokAdedi", "Kategori"]

# Diğer sütunları silmek
df = df[columns_to_keep]

# Düzenlenmiş dosyayı aynı adla kaydetmek
df.to_excel("Öne Çıkanlar.xlsx", index=False)

#endregion

#region Stok Adedini ETOPLA Yapma ve StokAdediToplam Sütununa Yerleştirme

# Öne Çıkanlar.xlsx dosyasını oku
df = pd.read_excel("Öne Çıkanlar.xlsx")

# UrunAdi sütununa göre gruplandırma yap ve StokAdedi sütununu topla
stok_adedi_toplam = df.groupby('UrunAdi')['StokAdedi'].sum().reset_index()

# Yeni bir StokAdediToplam sütunu ekleyelim
df = df.merge(stok_adedi_toplam, on='UrunAdi', how='left', suffixes=('', '_Toplam'))

# Yeni oluşturduğumuz StokAdediToplam sütununu adlandıralım
df.rename(columns={'StokAdedi_Toplam': 'StokAdediToplam'}, inplace=True)

# Sonucu yeni bir Excel dosyasına kaydedelim
df.to_excel("Öne Çıkanlar.xlsx", index=False)

#endregion

#region Beden Durumunu Hesaplama

# Veriyi Okuma
df = pd.read_excel('Öne Çıkanlar.xlsx')

# "StokAdedi" Sütununda 0'dan Büyük Olan Değerlerin Adedi
df['StokAdedi_GT_0'] = df['StokAdedi'].apply(lambda x: 1 if x > 0 else 0)
stok_adedi_gt_0_adet = df.groupby('UrunAdi')['StokAdedi_GT_0'].sum().reset_index()

# "UrunAdi" Sütunundaki Toplam Yenilenme Adedi
toplam_yenilenme_adedi = df.groupby('UrunAdi').size().reset_index(name='ToplamYenilenmeAdedi')

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0'], axis=1, errors='ignore')


# Oranı Hesapla ve Yeni Sütunu Ekle
df = pd.merge(df, stok_adedi_gt_0_adet, on='UrunAdi', how='left')
df = pd.merge(df, toplam_yenilenme_adedi, on='UrunAdi', how='left')
df['Beden Durumu'] = df['StokAdedi_GT_0'] / df['ToplamYenilenmeAdedi']

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0', 'ToplamYenilenmeAdedi'], axis=1, errors='ignore')

# Oranı 100 ile Çarpma
df['Beden Durumu'] *= 100

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('Öne Çıkanlar.xlsx', index=False)

#endregion

#region Beden Durumu %51'nin Altında Olanları Silme

# Excel dosyasını oku
birlesmis_veri = pd.read_excel("Öne Çıkanlar.xlsx")

# "Beden Durumu" sütunundaki değeri 50'den küçük olan satırları filtrele
birlesmis_veri = birlesmis_veri[birlesmis_veri["Beden Durumu"] >= 51]

# İstenmeyen sütunları sil
silinecek_sutunlar = ["Beden Durumu", "StokAdedi"]
birlesmis_veri = birlesmis_veri.drop(columns=silinecek_sutunlar, errors='ignore')

# Tekrar eden satırları kaldır
birlesmis_veri = birlesmis_veri.drop_duplicates()

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("Öne Çıkanlar.xlsx", index=False)

#endregion

#region Öne Çıkanlar Kategorisinden İndirimli Ürünleri Hariç Tutma

# Excel dosyasını oku
df = pd.read_excel('Öne Çıkanlar.xlsx')

# "Kategori" sütununda "İNDİRİM" ifadesini içeren satırları filtrele
df_filtered = df[~df['Kategori'].str.contains('İndirim', na=False)]

# Sonuçları aynı Excel dosyasına kaydet
df_filtered.to_excel('Öne Çıkanlar.xlsx', index=False)

#endregion

#region Satış Raporu Tarihini Düne Göre Ayarlama

# Excel dosyasının ismi ve konumu
filename = "Satış Raporu.xlsx"

# Dosyanın indirilme tarihini kontrol eden fonksiyon
def is_file_downloaded_today(file_path):
    if os.path.exists(file_path):
        # Dosyanın son değiştirilme tarihini al
        file_modification_time = os.path.getmtime(file_path)
        modification_date = datetime.fromtimestamp(file_modification_time).date()
        # Bugünün tarihi ile karşılaştır
        return modification_date == datetime.today().date()
    return False

# Eğer dosya bugün indirilmemişse Selenium işlemleri çalıştırılır
if not is_file_downloaded_today(filename):
    # ChromeDriver'ı en son sürümüyle otomatik olarak indirip kullan
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    login_url = "https://task.haydigiy.com/kullanici-giris/?ReturnUrl=%2Fadmin"
    driver.get(login_url)

    # Giriş bilgilerini doldurma
    email_input = driver.find_element("id", "EmailOrPhone")
    email_input.send_keys("mustafa_kod@haydigiy.com")

    password_input = driver.find_element("id", "Password")
    password_input.send_keys("123456")
    password_input.send_keys(Keys.RETURN)

    # Belirttiğiniz sayfaya yönlendirme
    desired_page_url = "https://task.haydigiy.com/admin/exportorder/edit/149/"
    driver.get(desired_page_url)

    # Dünün tarihini al
    yesterday = datetime.now() - timedelta(days=1)
    formatted_date = yesterday.strftime("%d.%m.%Y")

    # EndDate alanını bulma ve tarih girişini yapma
    end_date_input = driver.find_element("id", "EndDate")
    end_date_input.clear()  # Eğer mevcut bir değer varsa temizleyin
    end_date_input.send_keys(formatted_date)

    # StartDate alanını bulma ve tarih girişini yapma
    start_date_input = driver.find_element("id", "StartDate")
    start_date_input.clear()  # Eğer mevcut bir değer varsa temizleyin
    start_date_input.send_keys(formatted_date)

    # Kaydet butonunu bulma ve tıklama
    save_button = driver.find_element("css selector", 'button.btn.btn-primary[name="save"]')
    save_button.click()

    # Selenium işlemleri tamamlandıktan sonra tarayıcıyı kapatın
    driver.quit()

#endregion

#region Satış Raporunu İndirme

# Excel dosyasının indirileceği URL
url = "https://task.haydigiy.com/FaprikaOrderXls/Q7DC67/1/"
filename = "Satış Raporu.xlsx"

# Dosyanın indirilme tarihini kontrol etmek için fonksiyon
def is_file_downloaded_today(file_path):
    if os.path.exists(file_path):
        # Dosyanın son değiştirilme tarihini al
        file_modification_time = os.path.getmtime(file_path)
        modification_date = datetime.fromtimestamp(file_modification_time).date()
        # Bugünün tarihi ile karşılaştır
        return modification_date == datetime.today().date()
    return False

# Dosya bugün indirilmemişse veya yoksa yeniden indir
if not is_file_downloaded_today(filename):
    # Eğer dosya varsa sil
    if os.path.exists(filename):
        os.remove(filename)
    # Dosyayı indir ve kaydet
    response = requests.get(url)
    with open(filename, "wb") as file:
        file.write(response.content)

# Excel dosyasını oku
df = pd.read_excel(filename)

# Tutulacak sütunlar
columns_to_keep = ["UrunAdi", "Adet", "ToplamFiyat"]

# Diğer sütunları silmek
df = df[columns_to_keep]

# Düzenlenmiş dosyayı aynı adla kaydet
df.to_excel(filename, index=False)

#endregion

#region Adet Sütununu Sayıya Çevirme

def clean_adet(data):
    # "Adet" sütunundaki tüm verilerin virgül karakterinden sonrasını temizle
    data['Adet'] = data['Adet'].astype(str).apply(lambda x: x.split(',')[0])

if __name__ == "__main__":
    # Mevcut Excel dosyasını oku
    file_path = "Satış Raporu.xlsx"
    combined_data = pd.read_excel(file_path, engine="openpyxl")

    # "Adet" sütunundaki verilerin virgül karakterinden sonrasını temizle
    clean_adet(combined_data)

    # Güncellenmiş veriyi aynı dosyaya üstüne kaydet
    combined_data.to_excel(file_path, index=False, engine='openpyxl')
    
   
def convert_adet_to_numeric(data):
    # "Adet" sütunundaki tüm verileri sayıya dönüştür
    data['Adet'] = pd.to_numeric(data['Adet'], errors='coerce')  # 'coerce' ile hatalı değerleri NaN olarak işaretle

if __name__ == "__main__":
    # Mevcut Excel dosyasını oku
    file_path = "Satış Raporu.xlsx"
    combined_data = pd.read_excel(file_path, engine="openpyxl")

    # "Adet" sütunundaki verileri sayıya dönüştür
    convert_adet_to_numeric(combined_data)

    # Güncellenmiş veriyi aynı dosyaya üstüne kaydet
    combined_data.to_excel(file_path, index=False, engine='openpyxl')

#endregion

#region ToplamFiyat Sütununu Sayıya Çevirme

def clean_adet(data):
    # "Adet" sütunundaki tüm verilerin virgül karakterinden sonrasını temizle
    data['ToplamFiyat'] = data['ToplamFiyat'].astype(str).apply(lambda x: x.split(',')[0])

if __name__ == "__main__":
    # Mevcut Excel dosyasını oku
    file_path = "Satış Raporu.xlsx"
    combined_data = pd.read_excel(file_path, engine="openpyxl")

    # "Adet" sütunundaki verilerin virgül karakterinden sonrasını temizle
    clean_adet(combined_data)

    # Güncellenmiş veriyi aynı dosyaya üstüne kaydet
    combined_data.to_excel(file_path, index=False, engine='openpyxl')
    
   
def convert_adet_to_numeric(data):
    # "Adet" sütunundaki tüm verileri sayıya dönüştür
    data['ToplamFiyat'] = pd.to_numeric(data['ToplamFiyat'], errors='coerce')  # 'coerce' ile hatalı değerleri NaN olarak işaretle

if __name__ == "__main__":
    # Mevcut Excel dosyasını oku
    file_path = "Satış Raporu.xlsx"
    combined_data = pd.read_excel(file_path, engine="openpyxl")

    # "Adet" sütunundaki verileri sayıya dönüştür
    convert_adet_to_numeric(combined_data)

    # Güncellenmiş veriyi aynı dosyaya üstüne kaydet
    combined_data.to_excel(file_path, index=False, engine='openpyxl')

#endregion

#region Adet ve ToplamFiyat Sütununa ETOPLA yapma

# Excel dosyasını tekrar okumak
df = pd.read_excel("Satış Raporu.xlsx")

# UrunAdi sütununa göre gruplandırma ve Adet ile ToplamFiyat sütunlarındaki verileri toplama
df_grouped = df.groupby('UrunAdi', as_index=False).agg({
    'Adet': 'sum',
    'ToplamFiyat': 'sum'
})

# Düzenlenmiş dosyayı aynı adla kaydetmek
df_grouped.to_excel("Satış Raporu.xlsx", index=False)

#endregion

#region Öne Çıkanlar Listesine Satış Adetleri Listesindeki Verileri Çektirme

# Excel dosyalarını oku
satis_raporu_df = pd.read_excel("Satış Raporu.xlsx")
one_cikanlar_df = pd.read_excel("Öne Çıkanlar.xlsx")

# Öne Çıkanlar Excel'ine Satış Raporu'ndan Adet ve ToplamFiyat sütunlarını eklemek için merge işlemi yapalım
merged_df = one_cikanlar_df.merge(
    satis_raporu_df[['UrunAdi', 'Adet', 'ToplamFiyat']],
    on='UrunAdi',
    how='left'
)

# Birleştirilmiş veriyi Öne Çıkanlar Excel dosyasına kaydedelim
merged_df.to_excel("Öne Çıkanlar.xlsx", index=False)

#endregion

#region StokAdedi Satış Adedinden Büyük Olan Ürünleri Listeden Temizleme

# Öne Çıkanlar.xlsx dosyasını oku
df = pd.read_excel("Öne Çıkanlar.xlsx")

# StokAdediToplam sütunundaki veri Adet sütunundaki veriden küçükse satırı sil
df_filtered = df[df['StokAdediToplam'] >= df['Adet']]

# Sonucu yeni bir Excel dosyasına kaydedelim
df_filtered.to_excel("Öne Çıkanlar.xlsx", index=False)

#endregion

#region Ürünün Karlılığını Hesaplama

# Öne Çıkanlar Excel dosyasını oku
one_cikanlar_df = pd.read_excel("Öne Çıkanlar.xlsx")

# 'Adet' ve 'AlisFiyati' sütunlarındaki verileri çarp ve 'ToplamFiyat' sütunundan çıkar
one_cikanlar_df['ToplamFiyat'] = one_cikanlar_df['ToplamFiyat'] - (one_cikanlar_df['Adet'] * one_cikanlar_df['AlisFiyati'])

# Güncellenmiş veriyi Öne Çıkanlar Excel dosyasına kaydet
one_cikanlar_df.to_excel("Öne Çıkanlar.xlsx", index=False)

#endregion

#region Öne Çıkanlar Excel'inden Büyük Beden - Çocuk - Kolye - Küpe - Bileklik İçerenleri Silme

# Excel dosyasını okuyun
df = pd.read_excel('Öne Çıkanlar.xlsx')

# Silinmesi gereken kategoriler
kategoriler = ["Büyük Beden", "Çocuk", "Kolye", "Küpe", "Bileklik", "TESETTÜR"]

# Kategori sütununda belirtilen kategorileri içeren satırları filtrelemek için bir fonksiyon tanımlayın
def contains_any_category(cell_value):
    if pd.isna(cell_value):
        return False
    # Kategori listesinde herhangi birini hücre değerinde arayın
    return any(category in cell_value for category in kategoriler)

# Sadece belirtilen kategorileri içermeyen satırları filtreleyin
df_filtered = df[~df['Kategori'].apply(contains_any_category)]

# Filtrelenmiş DataFrame'i yeni bir Excel dosyasına yazın
df_filtered.to_excel('Öne Çıkanlar.xlsx', index=False)

#endregion

#region 150 Kar Eden ve 150 Çok Satan Ürünü Tespit Etme

# Öne Çıkanlar Excel dosyasını oku
df = pd.read_excel("Öne Çıkanlar.xlsx")

# 1. Adım: ToplamFiyat sütununda en büyük 150 değeri "Öne Çıkanlar Serbest Alan" olarak değiştir
top_75_toplam_fiyat = df.nlargest(150, 'ToplamFiyat')
df['ToplamFiyat'] = df.apply(lambda row: 'ÖNE ÇIKANLAR' if row['ToplamFiyat'] in top_75_toplam_fiyat['ToplamFiyat'].values else None, axis=1)

# 2. Adım: Adet sütununda en büyük 150 değeri, ToplamFiyat sütununda boş olan satırlara "Öne Çıkanlar Serbest Alan" olarak ata
top_75_adet = df[df['ToplamFiyat'].isna()].nlargest(150, 'Adet')
df.loc[df['ToplamFiyat'].isna() & df['UrunAdi'].isin(top_75_adet['UrunAdi']), 'ToplamFiyat'] = 'ÖNE ÇIKANLAR'

# Güncellenmiş veriyi Öne Çıkanlar Excel dosyasına kaydet
df.to_excel("Öne Çıkanlar.xlsx", index=False)

#endregion

#region Öne Çıkanlar Yükleme ve Öne Çıkanlar Excellerinin Arasında Veri Çektirme

# Excel dosyalarını oku
one_cikanlar = pd.read_excel("Öne Çıkanlar.xlsx")
one_cikanlar_yukleme = pd.read_excel("Öne Çıkanlar Yükleme.xlsx")

# "UrunAdi" sütununu kullanarak verileri eşleştir
# ToplamFiyat değerlerini Kategori sütununa çek
one_cikanlar_yukleme = one_cikanlar_yukleme.merge(
    one_cikanlar[['UrunAdi', 'ToplamFiyat']],
    on='UrunAdi',
    how='left'
)

# Kategori sütununu güncelle
one_cikanlar_yukleme['Kategori'] = one_cikanlar_yukleme['ToplamFiyat']

# Karşılığı bulunmayan satırları sil
one_cikanlar_yukleme = one_cikanlar_yukleme.dropna(subset=['Kategori'])

# İstenmeyen sütunları temizle
one_cikanlar_yukleme = one_cikanlar_yukleme.drop(columns=['ToplamFiyat'])

# Güncellenmiş veriyi Öne Çıkanlar Yükleme dosyasına kaydet
one_cikanlar_yukleme.to_excel("Öne Çıkanlar Yükleme.xlsx", index=False)

#endregion

#region Öne Çıkanlar Excel'ini Kopyalama ve Düzenleme

# Dosya adlarını belirle
orijinal_dosya = "Öne Çıkanlar Yükleme.xlsx"
yeni_dosya = "Öne Çıkanlar Sıralama.xlsx"

# Dosyayı kopyala
shutil.copy(orijinal_dosya, yeni_dosya)

# Yeni dosyayı oku
df = pd.read_excel(yeni_dosya)

# "UrunAdi" sütunu hariç diğer tüm sütunları sil
df = df[['UrunAdi']]

# Tekrar eden satırları kaldır
df = df.drop_duplicates()

# Güncellenmiş veriyi yeni dosyaya kaydet
df.to_excel(yeni_dosya, index=False)

#endregion

#region Öne Çıkanlar Sıralama Listesine Satış Adetleri Listesindeki Verileri Çektirme

# Excel dosyalarını oku
satis_raporu_df = pd.read_excel("Satış Raporu.xlsx")
one_cikanlar_df = pd.read_excel("Öne Çıkanlar Sıralama.xlsx")

# Öne Çıkanlar Excel'ine Satış Raporu'ndan Adet ve ToplamFiyat sütunlarını eklemek için merge işlemi yapalım
merged_df = one_cikanlar_df.merge(
    satis_raporu_df[['UrunAdi', 'Adet', 'ToplamFiyat']],
    on='UrunAdi',
    how='left'
)

# Birleştirilmiş veriyi Öne Çıkanlar Excel dosyasına kaydedelim
merged_df.to_excel("Öne Çıkanlar Sıralama.xlsx", index=False)

#endregion

#region Ürünlerin Sıralama Kurgusunu Ayarlama

# Excel dosyasını oku
dosya_adi = "Öne Çıkanlar Sıralama.xlsx"
xls = pd.ExcelFile(dosya_adi)

# Sheet1 sayfasını oku
sheet1_df = pd.read_excel(xls, sheet_name='Sheet1')

# DataFrame'leri tanımla
toplam_fiyat_df = sheet1_df[['UrunAdi', 'ToplamFiyat']].sort_values(by='ToplamFiyat', ascending=False).drop_duplicates()
adet_df = sheet1_df[['UrunAdi', 'Adet']].sort_values(by='Adet', ascending=False).drop_duplicates()

# En büyük değerler listesini oluştur
en_buyukler = []
alınan_urunler = set()

toplam_fiyat_indeks = 0
adet_indeks = 0

while toplam_fiyat_indeks < len(toplam_fiyat_df) or adet_indeks < len(adet_df):
    # ToplamFiyat'tan en büyük değeri al
    while toplam_fiyat_indeks < len(toplam_fiyat_df):
        urun_adi = toplam_fiyat_df.iloc[toplam_fiyat_indeks]['UrunAdi']
        if urun_adi not in alınan_urunler:
            en_buyukler.append(urun_adi)
            alınan_urunler.add(urun_adi)
            toplam_fiyat_indeks += 1
            break
        toplam_fiyat_indeks += 1

    # Adet'ten en büyük değeri al
    while adet_indeks < len(adet_df):
        urun_adi = adet_df.iloc[adet_indeks]['UrunAdi']
        if urun_adi not in alınan_urunler:
            en_buyukler.append(urun_adi)
            alınan_urunler.add(urun_adi)
            adet_indeks += 1
            break
        adet_indeks += 1

# Yeni sayfayı oluştur
with pd.ExcelWriter(dosya_adi, engine='openpyxl', mode='a') as writer:
    # Yeni sayfayı oluştur
    yeni_df = pd.DataFrame({'UrunAdi': en_buyukler})
    yeni_df.to_excel(writer, sheet_name='ÖNE ÇIKANLAR', index=False)

#endregion

#region Öne Çıkanlar Sıralama Excel'inde Sheet1 Sayfasını Silme

# Excel dosyasını oku
dosya_adi = "Öne Çıkanlar Sıralama.xlsx"

# Excel dosyasını aç
wb = load_workbook(dosya_adi)

# Sheet1 sayfasını sil
if 'Sheet1' in wb.sheetnames:
    del wb['Sheet1']

# Dosyayı kaydet
wb.save(dosya_adi)

#endregion

#region Öne Çıkanlar Sıralama Excel'inde Ürünlere Sıralama Verme

# Excel dosyasını oku
dosya_adi = "Öne Çıkanlar Sıralama.xlsx"
df = pd.read_excel(dosya_adi)

# UrunAdi sütunundaki hücre sayısını al
num_rows = len(df)

# Numara sütununu oluştur ve eksilterek doldur
df['Numara'] = range(-num_rows, 0)

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
df.to_excel(dosya_adi, index=False)

#endregion

#region Öne Çıkanlar Sıralama Excel'inde Ürünlere SayfaIsmi Verme

# Excel dosyasını oku
dosya_adi = "Öne Çıkanlar Sıralama.xlsx"
df = pd.read_excel(dosya_adi)

# SayfaIsmi sütununu oluştur ve tüm değerleri "ÖNE ÇIKANLAR" olarak doldur
df['SayfaIsmi'] = "ÖNE ÇIKANLAR"

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
df.to_excel(dosya_adi, index=False)

#endregion

#region Stabil Ürün Listesi Excel'ini Çoğaltma ve Adını Öne Çıkanlar Serbest Alan Yapma ve Serbest Alan Olmayanları Satır Olarak Silme

# Dosya adlarını tanımla
orijinal_dosya = "Stabil Ürün Listesi.xlsx"
kopya_dosya = "Öne Çıkanlar Serbest Alan.xlsx"

# Dosyayı kopyala
shutil.copy(orijinal_dosya, kopya_dosya)

# Kopyalanan dosyayı oku
df = pd.read_excel(kopya_dosya)

# "Kategori" sütununda "Öne Çıkanlar Serbest Alan" ifadesini içerenleri tut, diğerlerini sil
df_filtered = df[df['Kategori'].str.contains("Öne Çıkanlar Serbest Alan", na=False)]

# Filtrelenmiş veriyi yeni dosyaya kaydet
df_filtered.to_excel(kopya_dosya, index=False)

#endregion  

#region Öne Çıkanlar Serbest Alan Beden Durumunu Hesaplama

# Veriyi Okuma
df = pd.read_excel('Öne Çıkanlar Serbest Alan.xlsx')

# "StokAdedi" Sütununda 0'dan Büyük Olan Değerlerin Adedi
df['StokAdedi_GT_0'] = df['StokAdedi'].apply(lambda x: 1 if x > 0 else 0)
stok_adedi_gt_0_adet = df.groupby('UrunAdi')['StokAdedi_GT_0'].sum().reset_index()

# "UrunAdi" Sütunundaki Toplam Yenilenme Adedi
toplam_yenilenme_adedi = df.groupby('UrunAdi').size().reset_index(name='ToplamYenilenmeAdedi')

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0'], axis=1, errors='ignore')


# Oranı Hesapla ve Yeni Sütunu Ekle
df = pd.merge(df, stok_adedi_gt_0_adet, on='UrunAdi', how='left')
df = pd.merge(df, toplam_yenilenme_adedi, on='UrunAdi', how='left')
df['Beden Durumu'] = df['StokAdedi_GT_0'] / df['ToplamYenilenmeAdedi']

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0', 'ToplamYenilenmeAdedi'], axis=1, errors='ignore')

# Oranı 100 ile Çarpma
df['Beden Durumu'] *= 100

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('Öne Çıkanlar Serbest Alan.xlsx', index=False)

#endregion

#region Öne Çıkanlar Serbest Alan Excel'inde Sadece UrunAdi sütununu Tutma

# Excel dosyasını okumak
df = pd.read_excel("Öne Çıkanlar Serbest Alan.xlsx")

# Tutulacak sütunlar
columns_to_keep = ["UrunAdi"]

# Diğer sütunları silmek
df = df[columns_to_keep]

# Düzenlenmiş dosyayı aynı adla kaydetmek
df.to_excel("Öne Çıkanlar Serbest Alan.xlsx", index=False)

#endregion

#region Öne Çıkanlar Serbest Alan Ürünlerinin Kategorideki Sırasını Bulma

# Öne Çıkanlar Serbest Alan dosyasını oku
df = pd.read_excel("Öne Çıkanlar Serbest Alan.xlsx")

# Ürünlerin sıralanacağı bir liste oluştur
urun_adlari = df["UrunAdi"].tolist()

# Ürünleri aramak için ürün URL'sine GET isteği gönderme
product_url = "https://task.haydigiy.com/one-cikanlar-serbest-alan/"
response = requests.get(product_url)
soup = BeautifulSoup(response.text, "html.parser")

# Ürünleri bulma
product_items = soup.find_all(class_="product-item")

# Ürün numaralarını bulma
urun_numaralari = []
for urun_adi in urun_adlari:
    for idx, item in enumerate(product_items):
        if urun_adi in item.text:
            urun_numaralari.append((urun_adi, idx + 1))  # 1 tabanlı indeks
            break
    else:
        urun_numaralari.append((urun_adi, None))  # Bulunamazsa None

# Excel'e geri yazma
df['Numara'] = [numara for urun, numara in urun_numaralari]
df.to_excel("Öne Çıkanlar Serbest Alan.xlsx", index=False)

#endregion

#region Öne Çıkanlar Serbest Alan Ürünlerinin Sırasını Ayarlama

# Excel dosyasını oku
df = pd.read_excel("Öne Çıkanlar Serbest Alan.xlsx")

# Numara sütununa göre küçükten büyüğe sıralama
df_sorted = df.sort_values(by="Numara")

# Numara sütununu güncelle
df_sorted["Numara"] = range(-9999, -9999 + len(df_sorted), +1)

# Güncellenmiş veriyi Excel dosyasına kaydet
df_sorted.to_excel("Öne Çıkanlar Serbest Alan.xlsx", index=False)

#endregion

#region Öne Çıkanlar Serbest Alan Sıralama Excel'inde Ürünlere SayfaIsmi Verme

# Excel dosyasını oku
dosya_adi = "Öne Çıkanlar Serbest Alan.xlsx"
df = pd.read_excel(dosya_adi)

# SayfaIsmi sütununu oluştur ve tüm değerleri "ÖNE ÇIKANLAR" olarak doldur
df['SayfaIsmi'] = "ÖNE ÇIKANLAR"

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
df.to_excel(dosya_adi, index=False)

#endregion

#region Öne Çıkanlar Serbest Alan ve Öne Çıkanlar Sıralama Verilerini Alt Alta Birleştirme

# Excel dosyalarını oku
df_serbest_alan = pd.read_excel("Öne Çıkanlar Serbest Alan.xlsx")
df_siralama = pd.read_excel("Öne Çıkanlar Sıralama.xlsx")

# İki DataFrame'i birleştir
combined_df = pd.concat([df_siralama, df_serbest_alan], ignore_index=True)

# Birleştirilmiş veriyi yeni bir Excel dosyasına kaydet
combined_df.to_excel("Öne Çıkanlar Sıralama.xlsx", index=False)

#endregion

#region Stabil Ürün Listesi Excel'ini Çoğaltma ve Adını Sezon Sonu İndirimleri Yapma ve Sezon Sonu Olmayanları Satır Olarak Silme

# Dosya adlarını tanımla
orijinal_dosya = "Stabil Ürün Listesi.xlsx"
kopya_dosya = "Sezon Sonu İndirimleri.xlsx"

# Dosyayı kopyala
shutil.copy(orijinal_dosya, kopya_dosya)

# Kopyalanan dosyayı oku
df = pd.read_excel(kopya_dosya)

# "Kategori" sütununda "Dev İndirimler Serbest Alan" ifadesini içerenleri tut, diğerlerini sil
df_filtered = df[df['Kategori'].str.contains("İNDİRİMLİ ÜRÜNLER", na=False)]

# Filtrelenmiş veriyi yeni dosyaya kaydet
df_filtered.to_excel(kopya_dosya, index=False)

#endregion

#region Sezon Sonu İndirimleri Liste Fiyatı Ayarlama

# Veriyi Okuma
df = pd.read_excel('Sezon Sonu İndirimleri.xlsx')

# Alış Fiyatına Göre İşlemler ve Kategori Kontrolü
def calculate_list_price(row):
    alis_fiyati = row['AlisFiyati']
    kategori = row['Kategori']

    if 0 <= alis_fiyati <= 24.99:
        result = alis_fiyati + 10
    elif 25 <= alis_fiyati <= 39.99:
        result = alis_fiyati + 13
    elif 40 <= alis_fiyati <= 59.99:
        result = alis_fiyati + 17
    elif 60 <= alis_fiyati <= 200.99:
        result = alis_fiyati * 1.30
    elif alis_fiyati >= 201:
        result = alis_fiyati * 1.25
    else:
        result = alis_fiyati

    # KDV
    if isinstance(kategori, str) and any(category in kategori for category in ["Parfüm", "Gözlük", "Saat"]):
        result *= 1.20
    else:
        result *= 1.10

    return result


# Yeni Sütun Oluşturma
df['ListeFiyati'] = df.apply(calculate_list_price, axis=1)

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('Sezon Sonu İndirimleri.xlsx', index=False)

#endregion

#region Sezon Sonu İndirimleri İndirim Oranı Hesaplama

# Excel dosyasını oku
birlesmis_veri = pd.read_excel("Sezon Sonu İndirimleri.xlsx")

# "ListeFiyati" ve "SatisFiyati" sütunlarındaki verilerden işlem yap
birlesmis_veri["İndirimOrani"] = (birlesmis_veri["ListeFiyati"] - birlesmis_veri["SatisFiyati"]) * 100 / birlesmis_veri["ListeFiyati"]

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("Sezon Sonu İndirimleri.xlsx", index=False)

#endregion

#region Sezon Sonu İndirimleri İndirim Tutarı Hesaplama

# Excel dosyasını oku
birlesmis_veri = pd.read_excel("Sezon Sonu İndirimleri.xlsx")

# "ListeFiyati" ve "SatisFiyati" sütunlarındaki verilerden işlem yap
birlesmis_veri["İndirimTutari"] = (birlesmis_veri["ListeFiyati"] - birlesmis_veri["SatisFiyati"])

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("Sezon Sonu İndirimleri.xlsx", index=False)

#endregion

#region Sezon Sonu İndirimleri Excel'inde Stok Adedi Sütununu Silme

# Excel dosyasını okuma
df = pd.read_excel('Sezon Sonu İndirimleri.xlsx')

# "StokAdedi" sütununu silme
df = df.drop(columns=['StokAdedi'])

# Güncellenmiş veriyi Excel dosyasına kaydetme
df.to_excel('Sezon Sonu İndirimleri.xlsx', index=False)
#endregion

#region Sezon Sonu İndirimleri Excel'inde Yenilenen Kaldırma

# Excel dosyasını okuma
df = pd.read_excel('Sezon Sonu İndirimleri.xlsx')

# Tekrarlayan satırları temizleme
df_benzersiz = df.drop_duplicates()

# Yeni haliyle Excel dosyasını kaydetme
df_benzersiz.to_excel('Sezon Sonu İndirimleri.xlsx', index=False)

#endregion

#region Sezon Sonu İndirimlerini İndirim Oranına Göre Sıralama

# Excel dosyasını oku
df = pd.read_excel("Sezon Sonu İndirimleri.xlsx")

# İndirim tutarına göre azalan sıralama
sorted_by_discount_value = df.sort_values(by="İndirimTutari", ascending=False)

# İndirim oranına göre azalan sıralama
sorted_by_discount_rate = df.sort_values(by="İndirimOrani", ascending=False)

# Sıralama için boş bir liste oluştur
final_sorted_list = []
used_indices = set()  # Kullanılan satır indekslerini takip etmek için set

# İndirim tutarı ve indirim oranı sırayla eklenerek final listesi oluşturma
while len(used_indices) < len(df):
    if not sorted_by_discount_value.empty:
        index = sorted_by_discount_value.index[0]
        if index not in used_indices:
            final_sorted_list.append(sorted_by_discount_value.iloc[0])
            used_indices.add(index)
        sorted_by_discount_value = sorted_by_discount_value.iloc[1:]

    if not sorted_by_discount_rate.empty:
        index = sorted_by_discount_rate.index[0]
        if index not in used_indices:
            final_sorted_list.append(sorted_by_discount_rate.iloc[0])
            used_indices.add(index)
        sorted_by_discount_rate = sorted_by_discount_rate.iloc[1:]

# Listeyi DataFrame'e çevir
final_sorted_df = pd.DataFrame(final_sorted_list)

# Güncellenmiş veriyi Excel dosyasına kaydet
final_sorted_df.to_excel("Sezon Sonu İndirimleri.xlsx", index=False)

#endregion

#region Sezon Sonu İndirimleri Excel'inde Sadece UrunAdi Sütununu Tutma

# Excel dosyasını okumak
df = pd.read_excel("Sezon Sonu İndirimleri.xlsx")

# Tutulacak sütunlar
columns_to_keep = ["UrunAdi"]

# Diğer sütunları silmek
df = df[columns_to_keep]

# Düzenlenmiş dosyayı aynı adla kaydetmek
df.to_excel("Sezon Sonu İndirimleri.xlsx", index=False)

#endregion

#region Sezon Sonu İndirimleri Excel'inde Sıralama Verme

# Excel dosyasını okuma (dosyanın adını ve yolunu kendi dosyanızla değiştirin)
df = pd.read_excel('Sezon Sonu İndirimleri.xlsx')

# Toplam ürün sayısını alıyoruz (satır sayısı)
urun_sayisi = len(df)

# 'Numara' adında yeni bir sütun ekleyip, değerleri -1200, -1199, ... şeklinde dolduruyoruz
df['Numara'] = range(-urun_sayisi, 0)

# Yeni haliyle Excel dosyasını kaydetme
df.to_excel('Sezon Sonu İndirimleri.xlsx', index=False)


#endregion

#region Sezon Sonu İndirimleri Excel'inde Ürünlere SayfaIsmi Verme

# Excel dosyasını oku
dosya_adi = "Sezon Sonu İndirimleri.xlsx"
df = pd.read_excel(dosya_adi)

# SayfaIsmi sütununu oluştur ve tüm değerleri "ÖNE ÇIKANLAR" olarak doldur
df['SayfaIsmi'] = "İNDİRİMLİ ÜRÜNLER"

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
df.to_excel(dosya_adi, index=False)

#endregion

#region Sezon Sonu İndirimleri ve Öne Çıkanlar Sıralama Verilerini Alt Alta Birleştirme

# Excel dosyalarını oku
df_serbest_alan = pd.read_excel("Sezon Sonu İndirimleri.xlsx")
df_siralama = pd.read_excel("Öne Çıkanlar Sıralama.xlsx")

# İki DataFrame'i birleştir
combined_df = pd.concat([df_siralama, df_serbest_alan], ignore_index=True)

# Birleştirilmiş veriyi yeni bir Excel dosyasına kaydet
combined_df.to_excel("Öne Çıkanlar Sıralama.xlsx", index=False)

#endregion

#region Stabil Ürün Listesi Excel'ini Çoğaltma ve Adını Dev İndirimler Serbest Alan Yapma ve Serbest Alan Olmayanları Satır Olarak Silme

# Dosya adlarını tanımla
orijinal_dosya = "Stabil Ürün Listesi.xlsx"
kopya_dosya = "Dev İndirimler Serbest Alan.xlsx"

# Dosyayı kopyala
shutil.copy(orijinal_dosya, kopya_dosya)

# Kopyalanan dosyayı oku
df = pd.read_excel(kopya_dosya)

# "Kategori" sütununda "Dev İndirimler Serbest Alan" ifadesini içerenleri tut, diğerlerini sil
df_filtered = df[df['Kategori'].str.contains("Dev İndirimler Serbest Alan", na=False)]

# Filtrelenmiş veriyi yeni dosyaya kaydet
df_filtered.to_excel(kopya_dosya, index=False)

#endregion

#region Dev İndirimler Serbest Alan Beden Durumunu Hesaplama

# Veriyi Okuma
df = pd.read_excel('Dev İndirimler Serbest Alan.xlsx')

# "StokAdedi" Sütununda 0'dan Büyük Olan Değerlerin Adedi
df['StokAdedi_GT_0'] = df['StokAdedi'].apply(lambda x: 1 if x > 0 else 0)
stok_adedi_gt_0_adet = df.groupby('UrunAdi')['StokAdedi_GT_0'].sum().reset_index()

# "UrunAdi" Sütunundaki Toplam Yenilenme Adedi
toplam_yenilenme_adedi = df.groupby('UrunAdi').size().reset_index(name='ToplamYenilenmeAdedi')

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0'], axis=1, errors='ignore')


# Oranı Hesapla ve Yeni Sütunu Ekle
df = pd.merge(df, stok_adedi_gt_0_adet, on='UrunAdi', how='left')
df = pd.merge(df, toplam_yenilenme_adedi, on='UrunAdi', how='left')
df['Beden Durumu'] = df['StokAdedi_GT_0'] / df['ToplamYenilenmeAdedi']

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0', 'ToplamYenilenmeAdedi'], axis=1, errors='ignore')

# Oranı 100 ile Çarpma
df['Beden Durumu'] *= 100

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('Dev İndirimler Serbest Alan.xlsx', index=False)

#endregion

#region Dev İndirimler Serbest Alan Excel'inde Sadece UrunAdi Sütununu Tutma

# Excel dosyasını okumak
df = pd.read_excel("Dev İndirimler Serbest Alan.xlsx")

# Tutulacak sütunlar
columns_to_keep = ["UrunAdi"]

# Diğer sütunları silmek
df = df[columns_to_keep]

# Düzenlenmiş dosyayı aynı adla kaydetmek
df.to_excel("Dev İndirimler Serbest Alan.xlsx", index=False)

#endregion

#region Dev İndirimler Serbest Alan Ürünlerinin Kategorideki Sırasını Bulma

# Dev İndirimler Serbest Alan dosyasını oku
df = pd.read_excel("Dev İndirimler Serbest Alan.xlsx")

# Ürünlerin sıralanacağı bir liste oluştur
urun_adlari = df["UrunAdi"].tolist()

# Ürünleri aramak için ürün URL'sine GET isteği gönderme
product_url = "https://task.haydigiy.com/dev-indirimler-serbest-alan-2/"
response = requests.get(product_url)
soup = BeautifulSoup(response.text, "html.parser")

# Ürünleri bulma
product_items = soup.find_all(class_="product-item")

# Ürün numaralarını bulma
urun_numaralari = []
for urun_adi in urun_adlari:
    for idx, item in enumerate(product_items):
        if urun_adi in item.text:
            urun_numaralari.append((urun_adi, idx + 1))  # 1 tabanlı indeks
            break
    else:
        urun_numaralari.append((urun_adi, None))  # Bulunamazsa None

# Excel'e geri yazma
df['Numara'] = [numara for urun, numara in urun_numaralari]
df.to_excel("Dev İndirimler Serbest Alan.xlsx", index=False)

#endregion

#region Dev İndirimler Serbest Alan Ürünlerinin Sırasını Ayarlama

# Excel dosyasını oku
df = pd.read_excel("Dev İndirimler Serbest Alan.xlsx")

# Numara sütununa göre küçükten büyüğe sıralama
df_sorted = df.sort_values(by="Numara")

# Numara sütununu güncelle
df_sorted["Numara"] = range(-9999, -9999 + len(df_sorted), +1)

# Güncellenmiş veriyi Excel dosyasına kaydet
df_sorted.to_excel("Dev İndirimler Serbest Alan.xlsx", index=False)

#endregion

#region Dev İndirimler Serbest Alan Excel'inde Ürünlere SayfaIsmi Verme

# Excel dosyasını oku
dosya_adi = "Dev İndirimler Serbest Alan.xlsx"
df = pd.read_excel(dosya_adi)

# SayfaIsmi sütununu oluştur ve tüm değerleri "ÖNE ÇIKANLAR" olarak doldur
df['SayfaIsmi'] = "İNDİRİMLİ ÜRÜNLER"

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
df.to_excel(dosya_adi, index=False)

#endregion

#region Dev İndirimler Serbest Alan ve Öne Çıkanlar Sıralama Verilerini Alt Alta Birleştirme

# Excel dosyalarını oku
df_serbest_alan = pd.read_excel("Dev İndirimler Serbest Alan.xlsx")
df_siralama = pd.read_excel("Öne Çıkanlar Sıralama.xlsx")

# İki DataFrame'i birleştir
combined_df = pd.concat([df_siralama, df_serbest_alan], ignore_index=True)

# Birleştirilmiş veriyi yeni bir Excel dosyasına kaydet
combined_df.to_excel("Öne Çıkanlar Sıralama.xlsx", index=False)

#endregion

#region Stabil Ürün Listesi Excel'ini Çoğaltma ve Adını Yeni Gelenler Sıralama Yapma

# Dosya adlarını tanımla
orijinal_dosya = "Stabil Ürün Listesi.xlsx"
kopya_dosya = "Yeni Gelenler Sıralama.xlsx"

# Dosyayı kopyala
shutil.copy(orijinal_dosya, kopya_dosya)

# Kopyalanan dosyayı oku
df = pd.read_excel(kopya_dosya)

# "Kategori" sütununda "Dev İndirimler Serbest Alan" ifadesini içerenleri tut, diğerlerini sil
df_filtered = df[df['Kategori'].str.contains("YENİ GELENLER", na=False)]

# Filtrelenmiş veriyi yeni dosyaya kaydet
df_filtered.to_excel(kopya_dosya, index=False)

#endregion  

#region Yeni Gelenler Sıralama Beden Durumunu Hesaplama

# Veriyi Okuma
df = pd.read_excel('Yeni Gelenler Sıralama.xlsx')

# "StokAdedi" Sütununda 0'dan Büyük Olan Değerlerin Adedi
df['StokAdedi_GT_0'] = df['StokAdedi'].apply(lambda x: 1 if x > 0 else 0)
stok_adedi_gt_0_adet = df.groupby('UrunAdi')['StokAdedi_GT_0'].sum().reset_index()

# "UrunAdi" Sütunundaki Toplam Yenilenme Adedi
toplam_yenilenme_adedi = df.groupby('UrunAdi').size().reset_index(name='ToplamYenilenmeAdedi')

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0'], axis=1, errors='ignore')


# Oranı Hesapla ve Yeni Sütunu Ekle
df = pd.merge(df, stok_adedi_gt_0_adet, on='UrunAdi', how='left')
df = pd.merge(df, toplam_yenilenme_adedi, on='UrunAdi', how='left')
df['Beden Durumu'] = df['StokAdedi_GT_0'] / df['ToplamYenilenmeAdedi']

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0', 'ToplamYenilenmeAdedi'], axis=1, errors='ignore')

# Oranı 100 ile Çarpma
df['Beden Durumu'] *= 100

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('Yeni Gelenler Sıralama.xlsx', index=False)

#endregion

#region Yeni Gelenler Sıralama Beden Durumu %50'nin Altında Olanları Silme

# Excel dosyasını oku
birlesmis_veri = pd.read_excel("Yeni Gelenler Sıralama.xlsx")

# "Beden Durumu" sütunundaki değeri 50'den küçük olan satırları filtrele
birlesmis_veri = birlesmis_veri[birlesmis_veri["Beden Durumu"] >= 50]

# İstenmeyen sütunları sil
silinecek_sutunlar = ["Beden Durumu", "StokAdedi"]
birlesmis_veri = birlesmis_veri.drop(columns=silinecek_sutunlar, errors='ignore')

# Tekrar eden satırları kaldır
birlesmis_veri = birlesmis_veri.drop_duplicates()

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("Yeni Gelenler Sıralama.xlsx", index=False)

#endregion

#region Yeni Gelenler Sıralama Excel'inde Sadece UrunAdi ve Kategori Sütununu Tutma

# Excel dosyasını okumak
df = pd.read_excel("Yeni Gelenler Sıralama.xlsx")

# Tutulacak sütunlar
columns_to_keep = ["UrunAdi", "Kategori"]

# Diğer sütunları silmek
df = df[columns_to_keep]

# Düzenlenmiş dosyayı aynı adla kaydetmek
df.to_excel("Yeni Gelenler Sıralama.xlsx", index=False)

#endregion

#region Yeni Gelenler Sıralama Excel'inde Tekrar Stoktaları Tespit Etme

# Excel dosyasını oku
df = pd.read_excel('Yeni Gelenler Sıralama.xlsx')

# Kategori sütununu güncelle
df['Kategori'] = df['Kategori'].apply(lambda x: 'Tekrar Stokta' if 'Tekrar Stokta' in str(x) else '')

# Sonuçları Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('Yeni Gelenler Sıralama.xlsx', index=False)

#endregion

#region Stabil Ürün Listesi - Öne Çıkanlar Sıralama ve Yeni Gelenler Sıralama Excellerine Ürün ID'si Çektirme

# XML URL'si
xml_url = "https://task.haydigiy.com/FaprikaXml/KM283S/1/"

# XML'den Ürün Bilgilerini Çekme ve Temizleme
response = requests.get(xml_url)
xml_data = response.text
soup = BeautifulSoup(xml_data, 'xml')

product_data = []
for item in soup.find_all('item'):
    title = item.find('title').text.replace(' - Haydigiy', '')
    product_id = item.find('g:id').text if item.find('g:id') else None
    product_data.append({'UrunAdi': title, 'ID': product_id})

df_xml = pd.DataFrame(product_data)

# Stabil Ürün Listesi ile Birleştirme
df_calisma_alani = pd.read_excel('Stabil Ürün Listesi.xlsx')
df_merged_stabil = pd.merge(df_calisma_alani, df_xml, how='left', on='UrunAdi')
df_merged_stabil.to_excel('Stabil Ürün Listesi.xlsx', index=False)

# Öne Çıkanlar Sıralama ile Birleştirme
df_one_cikanlar = pd.read_excel('Öne Çıkanlar Sıralama.xlsx')
df_merged_one_cikanlar = pd.merge(df_one_cikanlar, df_xml, how='left', on='UrunAdi')
df_merged_one_cikanlar.to_excel('Öne Çıkanlar Sıralama.xlsx', index=False)

# Yeni Gelenler Sıralama ile Birleştirme
df_yeni_gelenler = pd.read_excel('Yeni Gelenler Sıralama.xlsx')
df_merged_yeni_gelenler = pd.merge(df_yeni_gelenler, df_xml, how='left', on='UrunAdi')
df_merged_yeni_gelenler.to_excel('Yeni Gelenler Sıralama.xlsx', index=False)

# Sezon Sonu Sıralama ile Birleştirme
df_sezon_sonu = pd.read_excel('Sezon Sonu İndirimleri.xlsx')
df_merged_sezon_sonu = pd.merge(df_yeni_gelenler, df_xml, how='left', on='UrunAdi')
df_merged_sezon_sonu.to_excel('Sezon Sonu İndirimleri.xlsx', index=False)

#endregion

#region Yeni Gelenler Sıralama Excel'i Sıralama Kurgusu

# Yeni Gelenler Sıralama excel dosyasını oku
df = pd.read_excel('Yeni Gelenler Sıralama.xlsx')

# Boş bir DataFrame oluştur ve "YENİ GELENLER" sayfasına ekleyeceğiz
df_new_sheet = pd.DataFrame(columns=['UrunAdi'])

# Kullanılmış UrunAdi'leri takip etmek için
used_urunadi = set()

# Tekrar Stokta verilerini sıralı şekilde almak
while True:
    # ID sütununa göre sıralama
    df_id_sorted = df.sort_values(by='ID', ascending=False)
    # Kategori sütununda "Tekrar Stokta" olanları seçme
    df_kategori_sorted = df[df['Kategori'] == 'Tekrar Stokta']
    
    # ID sütunundaki en büyük değeri al
    id_max_row = df_id_sorted[~df_id_sorted['UrunAdi'].isin(used_urunadi)].head(1)
    if not id_max_row.empty:
        df_new_sheet = pd.concat([df_new_sheet, pd.DataFrame({'UrunAdi': id_max_row['UrunAdi']})], ignore_index=True)
        used_urunadi.add(id_max_row['UrunAdi'].values[0])
    else:
        break
    
    # Kategori sütununda "Tekrar Stokta" olan en büyük değeri al
    if not df_kategori_sorted.empty:
        kategori_max_row = df_kategori_sorted[~df_kategori_sorted['UrunAdi'].isin(used_urunadi)].head(1)
        if not kategori_max_row.empty:
            df_new_sheet = pd.concat([df_new_sheet, pd.DataFrame({'UrunAdi': kategori_max_row['UrunAdi']})], ignore_index=True)
            used_urunadi.add(kategori_max_row['UrunAdi'].values[0])
        else:
            break

# Tekrar Stokta verileri bittiğinde kalan ID'lere göre ekleme
remaining_df = df[~df['UrunAdi'].isin(used_urunadi)]
remaining_sorted = remaining_df.sort_values(by='ID', ascending=False)
df_new_sheet = pd.concat([df_new_sheet, remaining_sorted[['UrunAdi']]], ignore_index=True)

# Yeni sayfayı ekle ve dosyayı kaydet
with pd.ExcelWriter('Yeni Gelenler Sıralama.xlsx', engine='openpyxl', mode='a') as writer:
    df_new_sheet.to_excel(writer, sheet_name='YENİ GELENLER', index=False)

#endregion

#region Yeni Gelenler Sıralama Excel'in de Sheet1 Sayfasını Silme

# Excel dosyasını oku
dosya_adi = "Yeni Gelenler Sıralama.xlsx"

# Excel dosyasını aç
wb = load_workbook(dosya_adi)

# Sheet1 sayfasını sil
if 'Sheet1' in wb.sheetnames:
    del wb['Sheet1']

# Dosyayı kaydet
wb.save(dosya_adi)

#endregion

#region Öne Çıkanlar Sıralama Excel'inde Ürünlere Sıralama Verme

# Excel dosyasını oku
dosya_adi = "Yeni Gelenler Sıralama.xlsx"
df = pd.read_excel(dosya_adi)

# UrunAdi sütunundaki hücre sayısını al
num_rows = len(df)

# Numara sütununu oluştur ve eksilterek doldur
df['Numara'] = range(-num_rows, 0)

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
df.to_excel(dosya_adi, index=False)

#endregion

#region Öne Çıkanlar Sıralama Excel'inde Ürünlere SayfaIsmi Verme

# Excel dosyasını oku
dosya_adi = "Yeni Gelenler Sıralama.xlsx"
df = pd.read_excel(dosya_adi)

# SayfaIsmi sütununu oluştur ve tüm değerleri "ÖNE ÇIKANLAR" olarak doldur
df['SayfaIsmi'] = "YENİ GELENLER"

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
df.to_excel(dosya_adi, index=False)

#endregion

#region Yeni Gelenler Sıralama Excel'ine Ürün ID'lerini Yeniden Çektirme


# Excel dosyalarını oku
df_yeni_gelenler = pd.read_excel('Yeni Gelenler Sıralama.xlsx', sheet_name='Sheet1')
df_stabil_urunler = pd.read_excel('Stabil Ürün Listesi.xlsx')

# ID'yi ilk karşılaştığı değeri baz alarak al
id_dict = df_stabil_urunler.drop_duplicates(subset='UrunAdi', keep='first').set_index('UrunAdi')['ID'].to_dict()

# Yeni Gelenler Sıralama'ya ID sütunu ekle
df_yeni_gelenler['ID'] = df_yeni_gelenler['UrunAdi'].map(id_dict)

# Yeni Gelenler Sıralama'ya ID sütununu ekleyip kaydet
with pd.ExcelWriter('Yeni Gelenler Sıralama.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_yeni_gelenler.to_excel(writer, sheet_name='Sheet1', index=False)

#endregion

#region Yeni Gelenler Sıralama ve Öne Çıkanlar Sıralama Verilerini Alt Alta Birleştirme

# Excel dosyalarını oku
df_serbest_alan = pd.read_excel("Yeni Gelenler Sıralama.xlsx")
df_siralama = pd.read_excel("Öne Çıkanlar Sıralama.xlsx")

# İki DataFrame'i birleştir
combined_df = pd.concat([df_siralama, df_serbest_alan], ignore_index=True)

# Birleştirilmiş veriyi yeni bir Excel dosyasına kaydet
combined_df.to_excel("Öne Çıkanlar Sıralama.xlsx", index=False)

#endregion

#region Stabil Ürün Listesi Excel'ini Çoğaltma ve Adını Yeni Sezon Serbest Alan Yapma ve Serbest Alan Olmayanları Satır Olarak Silme

# Dosya adlarını tanımla
orijinal_dosya = "Stabil Ürün Listesi.xlsx"
kopya_dosya = "Yeni Sezon Serbest Alan.xlsx"

# Dosyayı kopyala
shutil.copy(orijinal_dosya, kopya_dosya)

# Kopyalanan dosyayı oku
df = pd.read_excel(kopya_dosya)

# "Kategori" sütununda "Yeni Sezon Serbest Alan" ifadesini içerenleri tut, diğerlerini sil
df_filtered = df[df['Kategori'].str.contains("Yeni Sezon Serbest Alan", na=False)]

# Filtrelenmiş veriyi yeni dosyaya kaydet
df_filtered.to_excel(kopya_dosya, index=False)

#endregion

#region Yeni Sezon Serbest Alan Beden Durumunu Hesaplama

# Veriyi Okuma
df = pd.read_excel('Yeni Sezon Serbest Alan.xlsx')

# "StokAdedi" Sütununda 0'dan Büyük Olan Değerlerin Adedi
df['StokAdedi_GT_0'] = df['StokAdedi'].apply(lambda x: 1 if x > 0 else 0)
stok_adedi_gt_0_adet = df.groupby('UrunAdi')['StokAdedi_GT_0'].sum().reset_index()

# "UrunAdi" Sütunundaki Toplam Yenilenme Adedi
toplam_yenilenme_adedi = df.groupby('UrunAdi').size().reset_index(name='ToplamYenilenmeAdedi')

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0'], axis=1, errors='ignore')


# Oranı Hesapla ve Yeni Sütunu Ekle
df = pd.merge(df, stok_adedi_gt_0_adet, on='UrunAdi', how='left')
df = pd.merge(df, toplam_yenilenme_adedi, on='UrunAdi', how='left')
df['Beden Durumu'] = df['StokAdedi_GT_0'] / df['ToplamYenilenmeAdedi']

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0', 'ToplamYenilenmeAdedi'], axis=1, errors='ignore')

# Oranı 100 ile Çarpma
df['Beden Durumu'] *= 100

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('Yeni Sezon Serbest Alan.xlsx', index=False)

#endregion

#region Yeni Sezon Serbest Alan Excel'inde Sadece UrunAdi Sütununu Tutma1

# Excel dosyasını okumak
df = pd.read_excel("Yeni Sezon Serbest Alan.xlsx")

# Tutulacak sütunlar
columns_to_keep = ["UrunAdi"]

# Diğer sütunları silmek
df = df[columns_to_keep]

# Düzenlenmiş dosyayı aynı adla kaydetmek
df.to_excel("Yeni Sezon Serbest Alan.xlsx", index=False)

#endregion

#region Yeni Sezon Serbest Alan Ürünlerinin Kategorideki Sırasını Bulma

# Yeni Sezon Serbest Alan dosyasını oku
df = pd.read_excel("Yeni Sezon Serbest Alan.xlsx")

# Ürünlerin sıralanacağı bir liste oluştur
urun_adlari = df["UrunAdi"].tolist()

# Ürünleri aramak için ürün URL'sine GET isteği gönderme
product_url = "https://task.haydigiy.com/dev-indirimler-serbest-alan-2/"
response = requests.get(product_url)
soup = BeautifulSoup(response.text, "html.parser")

# Ürünleri bulma
product_items = soup.find_all(class_="product-item")

# Ürün numaralarını bulma
urun_numaralari = []
for urun_adi in urun_adlari:
    for idx, item in enumerate(product_items):
        if urun_adi in item.text:
            urun_numaralari.append((urun_adi, idx + 1))  # 1 tabanlı indeks
            break
    else:
        urun_numaralari.append((urun_adi, None))  # Bulunamazsa None

# Excel'e geri yazma
df['Numara'] = [numara for urun, numara in urun_numaralari]
df.to_excel("Yeni Sezon Serbest Alan.xlsx", index=False)

#endregion

#region Yeni Sezon Serbest Alan Ürünlerinin Sırasını Ayarlama

# Excel dosyasını oku
df = pd.read_excel("Yeni Sezon Serbest Alan.xlsx")

# Numara sütununa göre küçükten büyüğe sıralama
df_sorted = df.sort_values(by="Numara")

# Numara sütununu güncelle
df_sorted["Numara"] = range(-9999, -9999 + len(df_sorted), +1)

# Güncellenmiş veriyi Excel dosyasına kaydet
df_sorted.to_excel("Yeni Sezon Serbest Alan.xlsx", index=False)

#endregion

#region Yeni Sezon Serbest Alan Excel'inde Ürünlere SayfaIsmi Verme

# Excel dosyasını oku
dosya_adi = "Yeni Sezon Serbest Alan.xlsx"
df = pd.read_excel(dosya_adi)

# SayfaIsmi sütununu oluştur ve tüm değerleri "ÖNE ÇIKANLAR" olarak doldur
df['SayfaIsmi'] = "YENİ GELENLER"

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
df.to_excel(dosya_adi, index=False)

#endregion

#region Yeni Sezon Serbest Alan Excel'ine Ürün ID'lerini Yeniden Çektirme

# Excel dosyalarını oku
df_yeni_gelenler = pd.read_excel('Yeni Sezon Serbest ALan.xlsx', sheet_name='Sheet1')
df_stabil_urunler = pd.read_excel('Stabil Ürün Listesi.xlsx')

# ID'yi ilk karşılaştığı değeri baz alarak al
id_dict = df_stabil_urunler.drop_duplicates(subset='UrunAdi', keep='first').set_index('UrunAdi')['ID'].to_dict()

# Yeni Gelenler Sıralama'ya ID sütunu ekle
df_yeni_gelenler['ID'] = df_yeni_gelenler['UrunAdi'].map(id_dict)

# Yeni Gelenler Sıralama'ya ID sütununu ekleyip kaydet
with pd.ExcelWriter('Yeni Sezon Serbest Alan.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_yeni_gelenler.to_excel(writer, sheet_name='Sheet1', index=False)

#endregion

#region Yeni Sezon Serbest Alan ve Öne Çıkanlar Sıralama Verilerini Alt Alta Birleştirme

# Excel dosyalarını oku
df_serbest_alan = pd.read_excel("Yeni Sezon Serbest Alan.xlsx")
df_siralama = pd.read_excel("Öne Çıkanlar Sıralama.xlsx")

# İki DataFrame'i birleştir
combined_df = pd.concat([df_siralama, df_serbest_alan], ignore_index=True)

# Birleştirilmiş veriyi yeni bir Excel dosyasına kaydet
combined_df.to_excel("Öne Çıkanlar Sıralama.xlsx", index=False)

#endregion

#region Stabil Ürün Listesi Liste Fiyatı Ayarlama

# Veriyi Okuma
df = pd.read_excel('Stabil Ürün Listesi.xlsx')

# Alış Fiyatına Göre İşlemler ve Kategori Kontrolü
def calculate_list_price(row):
    alis_fiyati = row['AlisFiyati']
    kategori = row['Kategori']

    if 0 <= alis_fiyati <= 24.99:
        result = alis_fiyati + 10
    elif 25 <= alis_fiyati <= 39.99:
        result = alis_fiyati + 13
    elif 40 <= alis_fiyati <= 59.99:
        result = alis_fiyati + 17
    elif 60 <= alis_fiyati <= 200.99:
        result = alis_fiyati * 1.30
    elif alis_fiyati >= 201:
        result = alis_fiyati * 1.25
    else:
        result = alis_fiyati

    # KDV
    if isinstance(kategori, str) and any(category in kategori for category in ["Parfüm", "Gözlük", "Saat"]):
        result *= 1.20
    else:
        result *= 1.10

    return result


# Yeni Sütun Oluşturma
df['ListeFiyati'] = df.apply(calculate_list_price, axis=1)

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('Stabil Ürün Listesi.xlsx', index=False)

#endregion

#region Stabil Ürün Listesi Beden Durumunu Hesaplama

# Veriyi Okuma
df = pd.read_excel('Stabil Ürün Listesi.xlsx')

# "StokAdedi" Sütununda 0'dan Büyük Olan Değerlerin Adedi
df['StokAdedi_GT_0'] = df['StokAdedi'].apply(lambda x: 1 if x > 0 else 0)
stok_adedi_gt_0_adet = df.groupby('UrunAdi')['StokAdedi_GT_0'].sum().reset_index()

# "UrunAdi" Sütunundaki Toplam Yenilenme Adedi
toplam_yenilenme_adedi = df.groupby('UrunAdi').size().reset_index(name='ToplamYenilenmeAdedi')

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0'], axis=1, errors='ignore')


# Oranı Hesapla ve Yeni Sütunu Ekle
df = pd.merge(df, stok_adedi_gt_0_adet, on='UrunAdi', how='left')
df = pd.merge(df, toplam_yenilenme_adedi, on='UrunAdi', how='left')
df['Beden Durumu'] = df['StokAdedi_GT_0'] / df['ToplamYenilenmeAdedi']

# Gereksiz Sütunları Silme
df = df.drop(['StokAdedi_GT_0', 'ToplamYenilenmeAdedi'], axis=1, errors='ignore')

# Oranı 100 ile Çarpma
df['Beden Durumu'] *= 100

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('Stabil Ürün Listesi.xlsx', index=False)

#endregion

#region Stabil Ürün Listesi İndirim Oranı Hesaplama

# Excel dosyasını oku
birlesmis_veri = pd.read_excel("Stabil Ürün Listesi.xlsx")

# "ListeFiyati" ve "SatisFiyati" sütunlarındaki verilerden işlem yap
birlesmis_veri["İndirimOrani"] = (birlesmis_veri["ListeFiyati"] - birlesmis_veri["SatisFiyati"]) * 100 / birlesmis_veri["ListeFiyati"]

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("Stabil Ürün Listesi.xlsx", index=False)

#endregion

#region Stabil Ürün Listesi Beden Durumu %50'nin Altında Olanları Silme

# Excel dosyasını oku
birlesmis_veri = pd.read_excel("Stabil Ürün Listesi.xlsx")

# "Beden Durumu" sütunundaki değeri 50'den küçük olan satırları filtrele
birlesmis_veri = birlesmis_veri[birlesmis_veri["Beden Durumu"] >= 50]

# İstenmeyen sütunları sil
silinecek_sutunlar = ["Beden Durumu", "StokAdedi"]
birlesmis_veri = birlesmis_veri.drop(columns=silinecek_sutunlar, errors='ignore')

# Tekrar eden satırları kaldır
birlesmis_veri = birlesmis_veri.drop_duplicates()

# Veriyi mevcut Excel dosyasına kaydet (üzerine yaz)
birlesmis_veri.to_excel("Stabil Ürün Listesi.xlsx", index=False)

#endregion

#region Görüntülenmenin Satışa Dönüş Oranını Bulma


# Excel dosyasını oku
df = pd.read_excel('Stabil Ürün Listesi.xlsx')

# "Görüntülenmenin Satışa Dönüş Oranı" adında bir sütun oluştur ve hesaplamayı yap
df['Görüntülenmenin Satışa Dönüş Oranı'] = df['MorhipoKodu'] / df['HepsiBuradaKodu']

# Sonuçları aynı Excel dosyasına kaydet
df.to_excel('Stabil Ürün Listesi.xlsx', index=False)

#endregion

#region Stabil Ürün Listesindeki Gereksiz Sütunları Silme

# Excel dosyasını oku
df = pd.read_excel('Stabil Ürün Listesi.xlsx')

# "MorhipoKodu" ve "HepsiBuradaKodu" sütunlarını sil
df = df.drop(columns=['MorhipoKodu', 'HepsiBuradaKodu', 'AlisFiyati', 'SatisFiyati', 'ListeFiyati'])

# Sonuçları aynı Excel dosyasına kaydet
df.to_excel('Stabil Ürün Listesi.xlsx', index=False)

#endregion

#region Stabil Ürün Listesinde Tekrar Stoktaları Tespit Etme

# Excel dosyasını oku
df = pd.read_excel('Stabil Ürün Listesi.xlsx')

# "Tekrar Stokta" adında yeni bir sütun oluştur ve "Kategori" sütununda "Tekrar Stokta" içeren hücreler için yeni sütuna "Tekrar Stokta" yaz
df['Tekrar Stokta'] = df['Kategori'].apply(lambda x: 'Tekrar Stokta' if 'Tekrar Stokta' in str(x) else '')

# Sonuçları aynı Excel dosyasına kaydet
df.to_excel('Stabil Ürün Listesi.xlsx', index=False)

#endregion

#region Görüntülenmenin Satışa Dönüş Oranı Sütununda inf Olanları Sıfır Olarak Değiştirme

# Excel dosyasını oku
df = pd.read_excel('Stabil Ürün Listesi.xlsx')

# "Görüntülenmenin Satışa Dönüş Oranı" sütununda "inf" olan hücreleri 0 ile değiştir
df['Görüntülenmenin Satışa Dönüş Oranı'] = df['Görüntülenmenin Satışa Dönüş Oranı'].replace([np.inf, -np.inf], 0)

# Sonuçları aynı Excel dosyasına kaydet
df.to_excel('Stabil Ürün Listesi.xlsx', index=False)

#endregion

#region AramaTerimleri ve TrendyolKodu Sütunundaki Tarihleri Tespit Edip Çıkarma ve Güne Çevirme

# Exceli Oku
df_calisma_alani = pd.read_excel('Stabil Ürün Listesi.xlsx')

# Tarihleri çıkarma fonksiyonu
def extract_date(text):
    date_pattern = r'(\d{1,2}\.\d{1,2}\.\d{4})'
    match = re.search(date_pattern, str(text))
    return match.group(1) if match else None

# Tarihleri güne çevirme fonksiyonu
def calculate_days_to_today(date_str):
    if date_str is None:
        return date_str  # Tarih yoksa değeri olduğu gibi bırak
    tarih = datetime.strptime(date_str, '%d.%m.%Y')
    bugun = datetime.today()
    uzaklik = (bugun - tarih).days
    return uzaklik

# "AramaTerimleri" sütunundaki tarihleri temizle ve günlere çevir
df_calisma_alani['AramaTerimleri'] = df_calisma_alani['AramaTerimleri'].apply(lambda x: calculate_days_to_today(extract_date(x)))

# "TrendyolKodu" sütunundaki tarihleri temizle ve günlere çevir
df_calisma_alani['TrendyolKodu'] = df_calisma_alani['TrendyolKodu'].apply(lambda x: calculate_days_to_today(extract_date(x)))

# Exceli Kaydet
df_calisma_alani.to_excel('Stabil Ürün Listesi.xlsx', index=False, sheet_name='Sheet1')

#endregion

#region Tabloyu Görüntülenmenin Satışa Dönüş Oranına Göre Büyükten Küçüğe Sıralama

# Excel dosyasını oku
df = pd.read_excel('Stabil Ürün Listesi.xlsx')

# "Görüntülenmenin Satışa Dönüş Oranı" sütununa göre büyükten küçüğe sıralama
df = df.sort_values(by='Görüntülenmenin Satışa Dönüş Oranı', ascending=False)

# Sonuçları aynı Excel dosyasına kaydet
df.to_excel('Stabil Ürün Listesi.xlsx', index=False)

#endregion

#region e-Tablodan Veriyi Çekme ve Sıralanacak Kategorileri Sayfalara Bölme

# Google Sheets URL
google_sheet_url = "https://docs.google.com/spreadsheets/d/1suzb1TJyZz1xCtUxs1QOP7dHeWTKWkzDul_TaAtb1mc/gviz/tq?tqx=out:csv"

try:
    # Google Sheets'ten veriyi oku
    google_df = pd.read_csv(google_sheet_url)
    
    # "Sıralanacak Kategoriler" sütunundaki verileri al
    istenilen_kategoriler = google_df['Sıralanacak Kategoriler'].tolist()
    
    # Excel dosyasını oku
    df_birlesmis_veri = pd.read_excel('Stabil Ürün Listesi.xlsx')
    
    # İstenilen kategorileri içeren satırları filtrele
    for kategori in istenilen_kategoriler:
        df_kategori = df_birlesmis_veri[df_birlesmis_veri['Kategori'].str.contains(kategori, case=False, na=False)]
        
        # Veriyi yeni bir sayfaya kaydet (excel dosyasının içine eklenir)
        with pd.ExcelWriter('Stabil Ürün Listesi.xlsx', engine='openpyxl', mode='a') as writer:
            df_kategori.to_excel(writer, sheet_name=f'{kategori}', index=False)
        
except requests.exceptions.RequestException as e:
    pass

#endregion

#region 1-10 ve Yeni Gelen Tekrar Stokta Arasındaki Ürünlerin Sırasını Ayarlama

# Yeni bir Excel dosyası oluştur
yeni_birlesmis_veri = pd.ExcelWriter("Stabil Ürün Listesi Sıralama.xlsx", engine='xlsxwriter')

# Orijinal Excel dosyasını oku
birlesmis_veri = pd.read_excel("Stabil Ürün Listesi.xlsx", sheet_name=None)

# Her sayfa için işlemi yap
for sheet_name, df in birlesmis_veri.items():
    # Eğer veri yoksa işlem yapma
    if df.empty:
        continue
    
    # Tüm verilerden ilk 10 ID'yi al
    ilk_10_id = df[['ID']].head(10)
    
    # Eğer ID'ler 10'dan azsa, eksik olanları ekle
    if len(ilk_10_id) < 10:
        # Diğer ID'leri ekle
        kalan_id = df[~df['ID'].isin(ilk_10_id['ID'])][['ID']].head(10 - len(ilk_10_id))
        ilk_10_id = pd.concat([ilk_10_id, kalan_id], ignore_index=True)

    # Yeni bir veri çerçevesi oluştur ve ilk 10 ID'yi ekle
    yeni_df = pd.DataFrame({'ID': ilk_10_id['ID']})

    # AramaTerimleri ve TrendyolKodu sütunlarındaki uygun ID'leri bul
    arama_terimleri_filtered = df[(df['AramaTerimleri'].apply(lambda x: isinstance(x, (int, float)))) & 
                                  (df['AramaTerimleri'] >= 1) & (df['AramaTerimleri'] <= 3)]
    trendyol_kodu_filtered = df[(df['TrendyolKodu'].apply(lambda x: isinstance(x, (int, float)))) & 
                                (df['TrendyolKodu'] >= 1) & (df['TrendyolKodu'] <= 3)]
    
    # İndeksleri sıfırla
    arama_terimleri_filtered = arama_terimleri_filtered.sort_values(by='AramaTerimleri').reset_index(drop=True)
    trendyol_kodu_filtered = trendyol_kodu_filtered.sort_values(by='TrendyolKodu').reset_index(drop=True)

    # Döngü ile bir 'AramaTerimleri' ve bir 'TrendyolKodu' değerlerini al
    for i in range(max(len(arama_terimleri_filtered), len(trendyol_kodu_filtered))):
        # AramaTerimleri sütunundaki en küçük sayısal değerin ID'sini al
        if i < len(arama_terimleri_filtered):
            arama_id = arama_terimleri_filtered.loc[i, 'ID']
            if arama_id not in ilk_10_id['ID'].tolist() and arama_id not in yeni_df['ID'].tolist():
                yeni_df = pd.concat([yeni_df, pd.DataFrame({'ID': [arama_id]})], ignore_index=True)
        
        # TrendyolKodu sütunundaki en küçük sayısal değerin ID'sini al
        if i < len(trendyol_kodu_filtered):
            trendyol_id = trendyol_kodu_filtered.loc[i, 'ID']
            if trendyol_id not in ilk_10_id['ID'].tolist() and trendyol_id not in yeni_df['ID'].tolist():
                yeni_df = pd.concat([yeni_df, pd.DataFrame({'ID': [trendyol_id]})], ignore_index=True)

    # Yeni sayfayı oluştur ve veri çerçevesini yaz
    yeni_df.to_excel(yeni_birlesmis_veri, sheet_name=sheet_name, index=False)

# Yeni Excel dosyasını kaydet
yeni_birlesmis_veri.close()

#endregion

#region Stabil Ürün Listesi Excel'inde Kullanılan ID'leri Temizleme

# Stabil Ürün Listesi Sıralama dosyasındaki ID'leri oku
siralama_df = pd.read_excel("Stabil Ürün Listesi Sıralama.xlsx", sheet_name=None)

# Stabil Ürün Listesi dosyasını oku
birlesmis_df = pd.read_excel("Stabil Ürün Listesi.xlsx", sheet_name=None)

# Güncellenmiş dosyayı oluştur (varsa üzerine yazar)
with pd.ExcelWriter("Stabil Ürün Listesi.xlsx", engine='openpyxl') as writer:
    # Stabil Ürün Listesi Sıralama dosyasındaki her sayfa için işlemi yap
    for sheet_name in siralama_df.keys():
        if sheet_name in birlesmis_df:
            # Stabil Ürün Listesi Sıralama sayfasındaki ID'leri al
            siralama_id_list = siralama_df[sheet_name]['ID'].tolist()

            # Stabil Ürün Listesi sayfasını oku
            df = birlesmis_df[sheet_name]

            # Eğer veri yoksa işlem yapma
            if df.empty:
                continue

            # ID'lere göre filtreleme yaparak belirtilen ID'leri kaldır
            df_filtered = df[~df['ID'].isin(siralama_id_list)]

            # Filtrelenmiş veriyi aynı sayfa adıyla güncellenmiş dosyaya yaz
            df_filtered.to_excel(writer, sheet_name=sheet_name, index=False)

#endregion

#region 20-Sonsuz Ürünlerin Sırasını Ayarlama

# Yeni bir Excel dosyası oluştur
yeni_birlesmis_veri = pd.ExcelWriter("Stabil Ürün Listesi Sıralama2.xlsx", engine='xlsxwriter')

# Orijinal Excel dosyasını oku
birlesmis_veri = pd.read_excel("Stabil Ürün Listesi.xlsx", sheet_name=None)

# Her sayfa için işlemi yap
for sheet_name, df in birlesmis_veri.items():
    # Eğer veri yoksa işlem yapma
    if df.empty:
        continue

    # Sütun adlarının doğru olduğunu kontrol et
    required_columns = ['Görüntülenmenin Satışa Dönüş Oranı', 'ID']
    if not all(col in df.columns for col in required_columns):
        print(f"Uyarı: '{sheet_name}' sayfasında gerekli sütunlar eksik.")
        continue

    # 'Görüntülenmenin Satışa Dönüş Oranı' sütunundaki NaN değerleri temizle
    df = df.dropna(subset=['Görüntülenmenin Satışa Dönüş Oranı'])

    # 'Görüntülenmenin Satışa Dönüş Oranı' sütunundaki en yüksek değerlere sahip tüm ID'leri al
    sorted_df = df.sort_values(by='Görüntülenmenin Satışa Dönüş Oranı', ascending=False)
    
    # Tüm ID'leri al
    id_list = sorted_df['ID'].tolist()

    # ID listesini DataFrame'e dönüştür
    id_df = pd.DataFrame(id_list, columns=['ID'])

    # Yeni Excel dosyasına yaz
    id_df.to_excel(yeni_birlesmis_veri, sheet_name=sheet_name, index=False)

# Excel dosyasını kaydet
yeni_birlesmis_veri.close()

#endregion

#region 1-20 ile 20-Sonsuz Listelerini Birleştirme

# Eski Excel dosyasını ve yeni oluşturulan dosyayı oku
eski_dosya_yolu = "Stabil Ürün Listesi Sıralama.xlsx"
yeni_dosya_yolu = "Stabil Ürün Listesi Sıralama2.xlsx"
son_dosya_yolu = "Son Liste.xlsx"

# Eski dosyayı oku
with pd.ExcelFile(eski_dosya_yolu) as eski_dosya:
    # Yeni dosyayı oku
    with pd.ExcelFile(yeni_dosya_yolu) as yeni_dosya:
        # Yeni Excel dosyasını yazmak için bir ExcelWriter oluştur
        with pd.ExcelWriter(son_dosya_yolu, engine='xlsxwriter') as writer:
            # Eski dosyadaki her sayfayı oku ve yeni dosyadan eklemeleri yap
            for sheet_name in eski_dosya.sheet_names:
                # Eski dosyadaki veriyi oku
                eski_df = pd.read_excel(eski_dosya, sheet_name=sheet_name)
                
                # Yeni dosyadaki veriyi oku, eğer varsa
                if sheet_name in yeni_dosya.sheet_names:
                    yeni_df = pd.read_excel(yeni_dosya, sheet_name=sheet_name)
                    # Eski ve yeni veriyi birleştir
                    birlesmis_df = pd.concat([eski_df, yeni_df], ignore_index=True)
                else:
                    # Yeni dosyada bu sayfa yoksa eski veriyi kullan
                    birlesmis_df = eski_df

                # Birleşmiş veriyi yeni dosyaya yaz
                birlesmis_df.to_excel(writer, sheet_name=sheet_name, index=False)

#endregion

#region Son Liste Excel'inde Sheet1 Sayfasını Silme

# Excel dosyasını oku
df = pd.read_excel('Son Liste.xlsx', sheet_name=None)

# "Sheet1" sayfasını sil
df.pop('Sheet1', None)

# Yeni bir Excel dosyası olarak kaydet
with pd.ExcelWriter('Son Liste.xlsx', engine='openpyxl') as writer:
    for sheet_name, df_sheet in df.items():
        df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

#endregion

#region Son Liste Excel'inde SayfaIsmi Verme

# "Birlestirilmis_Veriler_Sirali" Excel dosyasını oku
df_birlestirilmis_sirali = pd.read_excel('Son Liste.xlsx', sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter('Son Liste.xlsx', engine='openpyxl') as writer:
    # Her bir sayfa için işlemleri gerçekleştir
    for sheet_name, df_sheet_birlestirilmis_sirali in df_birlestirilmis_sirali.items():
        # Sayfa ismini ekleyerek veriyi güncelle
        df_sheet_birlestirilmis_sirali['SayfaIsmi'] = sheet_name

        # Veriyi güncellenmiş haliyle aynı sayfaya kaydet
        df_sheet_birlestirilmis_sirali.to_excel(writer, sheet_name=sheet_name, index=False)

#endregion

#region Son Liste Excel'inde Sıralama Verme

# "Birlestirilmis_Veriler_Sirali_Yeni" Excel dosyasını oku
df_birlestirilmis_sirali_yeni = pd.read_excel('Son Liste.xlsx', sheet_name=None)

# Yeni bir Excel dosyası oluştur
with pd.ExcelWriter('Son Liste.xlsx', engine='openpyxl') as writer:
    # Her bir sayfa için işlemleri gerçekleştir
    for sheet_name, df_sheet_birlestirilmis_sirali_yeni in df_birlestirilmis_sirali_yeni.items():
        # Her satıra sırasıyla -50, -49, -48, ... şeklinde sayıları ekle
        df_sheet_birlestirilmis_sirali_yeni['Numara'] = range(-len(df_sheet_birlestirilmis_sirali_yeni), 0)

        # Veriyi güncellenmiş haliyle aynı sayfaya kaydet
        df_sheet_birlestirilmis_sirali_yeni.to_excel(writer, sheet_name=sheet_name, index=False)


#endregion

#region Son Liste Excel'inde Sayfaları Alt Alta Yerleştirme

# Excel dosyasının adı ve konumu
excel_dosyasi = "Son Liste.xlsx"

# Excel dosyasını yükle
birlesmis_veri = pd.ExcelFile(excel_dosyasi)

# Tüm sayfaların verilerini birleştirmek için boş bir DataFrame oluştur
birlesmis_df = pd.DataFrame()

# Her bir sayfa için işlem yap
for sayfa in birlesmis_veri.sheet_names:
    # Sayfa verisini oku
    veri = birlesmis_veri.parse(sayfa)
    
    # Her sayfanın verisini birleştir
    birlesmis_df = pd.concat([birlesmis_df, veri])

# Yeni bir dosya adı
yeni_dosya_adı = "Son Liste.xlsx"

# Birleştirilmiş veriyi yeni bir Excel dosyası olarak kaydet
birlesmis_df.to_excel(yeni_dosya_adı, index=False)

#endregion

#region Son Liste Excel'inde Sütunların Yerini Ayarlama

# Excel dosyasını oku
dosya_adı = "Son Liste.xlsx"
df = pd.read_excel(dosya_adı)

# Yeni sütun sıralamasını oluştur
yeni_sutun_sirasi = ['Numara', 'SayfaIsmi', 'ID']

# Sütunların sıralamasını yeniden düzenle
df = df[yeni_sutun_sirasi]

# Güncellenmiş veriyi yeni bir dosyaya kaydet
yeni_dosya_adı = "Son Liste.xlsx"
df.to_excel(yeni_dosya_adı, index=False)

#endregion

#region Öne Çıkanlar Sıralama Excel'inde UrunAdi Sütununu Silme

# Excel dosyasını oku
dosya_adı = "Öne Çıkanlar Sıralama.xlsx"
df = pd.read_excel(dosya_adı)

# 'UrunAdi' sütununu sil
if 'UrunAdi' in df.columns:
    df = df.drop(columns=['UrunAdi'])
else:
    pass

# Güncellenmiş veriyi yeni bir dosyaya kaydet
yeni_dosya_adı = "Öne Çıkanlar Sıralama.xlsx"
df.to_excel(yeni_dosya_adı, index=False)

#endregion

#region Son Liste Excel'i ile Öne Çıkanlar Sıralama Excel'ini Birleştirme

# Excel dosyalarını oku
son_liste_df = pd.read_excel("Son Liste.xlsx")
one_cikanlar_df = pd.read_excel("Öne Çıkanlar Sıralama.xlsx")

# (Varsayımsal olarak sütun isimlerini Son Liste ile uyumlu hale getiriyoruz, gerekirse düzenleme yapılabilir)
one_cikanlar_df = one_cikanlar_df[['Numara', 'SayfaIsmi', 'ID']]

# Verileri birleştir
birlesmis_df = pd.concat([son_liste_df, one_cikanlar_df], ignore_index=True)

# Güncellenmiş veriyi yeni bir dosyaya kaydet
yeni_dosya_adı = "Son Liste.xlsx"
birlesmis_df.to_excel(yeni_dosya_adı, index=False)

#endregion

#region Son Liste Düzenleme

# Google Sheets'ten CSV formatında veriyi oku
google_sheet_url = "https://docs.google.com/spreadsheets/d/1suzb1TJyZz1xCtUxs1QOP7dHeWTKWkzDul_TaAtb1mc/gviz/tq?tqx=out:csv"
google_df = pd.read_csv(google_sheet_url)

# "Son Liste.xlsx" dosyasını oku
son_liste_df = pd.read_excel("Son Liste.xlsx")

# Google Sheet'teki "Sıralanacak Kategoriler" ve "Kategori ID" sütunlarını al
kategoriler = google_df[['Sıralanacak Kategoriler', 'Kategori ID']]

# Kategori ID'lerini belirli koşullara göre ayarla
def kategori_id_bul(sayfa_isim):
    if sayfa_isim == "YENİ GELENLER":
        return 394
    elif sayfa_isim == "İNDİRİMLİ ÜRÜNLER":
        return 374
    elif sayfa_isim == "ÖNE ÇIKANLAR":
        return 26
    else:
        # Eğer bu isimlerden değilse, Google Sheet'teki verilerle eşleşme yap
        kategori = kategoriler[kategoriler['Sıralanacak Kategoriler'] == sayfa_isim]
        if not kategori.empty:
            return kategori.iloc[0]['Kategori ID']
        else:
            return None  # Eşleşme bulunamazsa

# "SayfaIsmi" sütunundaki her bir satır için uygun kategori ID'yi bul ve yaz
son_liste_df['Kategori ID'] = son_liste_df['SayfaIsmi'].apply(kategori_id_bul)

# Güncellenmiş veriyi yeni bir Excel dosyasına yaz
son_liste_df.to_excel("Son Liste.xlsx", index=False)

#endregion

#region Selenium Giriş Yapma

# ChromeDriver'ı en son sürümüyle otomatik olarak indirip kullan
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

login_url = "https://task.haydigiy.com/kullanici-giris/?ReturnUrl=%2Fadmin"
driver.get(login_url)

email_input = driver.find_element("id", "EmailOrPhone")
email_input.send_keys("mustafa_kod@haydigiy.com")

password_input = driver.find_element("id", "Password")
password_input.send_keys("123456")
password_input.send_keys(Keys.RETURN)

#endregion

#region Öne Çıkanlar Kategorisi Boşaltma

# Giriş yaptıktan sonra belirtilen sayfaya git
desired_page_url = "https://task.haydigiy.com/admin/product/bulkedit/"
driver.get(desired_page_url)

#Kategori Dahil Alan
category_select = Select(driver.find_element("id", "SearchInCategoryIds"))

#Kategori ID'si
category_select.select_by_value("26")

#Seçiniz Kısmının Tikini Kaldırma
all_remove_buttons = driver.find_elements(By.XPATH, "//span[@class='select2-selection__choice__remove']")
if len(all_remove_buttons) > 1:
    second_remove_button = all_remove_buttons[1]
    second_remove_button.click()
else:
    pass

#Ara Butonuna Tıklama
search_button = driver.find_element(By.ID, "search-products")
search_button.click()

# Sayfanın en sonuna git
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

# Kategori Güncelleme Tikine Tıklama
checkbox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "Category_Update")))
driver.execute_script("arguments[0].click();", checkbox)

# Kategori Güncelleme Alanını Bulma
category_id_select = driver.find_element(By.ID, "CategoryId")

# Kategori Güncelleme Alanında Kategori ID'si Seçme
category_id_select = Select(category_id_select)
category_id_select.select_by_value("26")

# Kategori Güncelleme Alanında Yapılacak İşlem Alanını Bulma
category_transaction_select = driver.find_element(By.ID, "CategoryTransactionId")

# Kategori Güncelleme Alanında Yapılacak İşlemi Seçme
category_transaction_select = Select(category_transaction_select)
category_transaction_select.select_by_value("1")

try:
    #Kaydet Butonunu Bulma
    save_button = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, 'btn-primary'))
    )

    #Kaydet Butonununa Tıklama
    driver.execute_script("arguments[0].click();", save_button)
except Exception as e:
    print(f"Hata: {e}")


#endregion

#region Öne Çıkanlar Kategorisine Öne Çıkanlar Serbest Alan Kategorisindeki Ürünleri Ekleme

# Giriş yaptıktan sonra belirtilen sayfaya git
desired_page_url = "https://task.haydigiy.com/admin/product/bulkedit/"
driver.get(desired_page_url)

#Kategori Dahil Alan
category_select = Select(driver.find_element("id", "SearchInCategoryIds"))

#Kategori ID'si
category_select.select_by_value("526")

#Seçiniz Kısmının Tikini Kaldırma
all_remove_buttons = driver.find_elements(By.XPATH, "//span[@class='select2-selection__choice__remove']")
if len(all_remove_buttons) > 1:
    second_remove_button = all_remove_buttons[1]
    second_remove_button.click()
else:
    pass

#Ara Butonuna Tıklama
search_button = driver.find_element(By.ID, "search-products")
search_button.click()

# Sayfanın en sonuna git
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

# Kategori Güncelleme Tikine Tıklama
checkbox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "Category_Update")))
driver.execute_script("arguments[0].click();", checkbox)

# Kategori Güncelleme Alanını Bulma
category_id_select = driver.find_element(By.ID, "CategoryId")

# Kategori Güncelleme Alanında Kategori ID'si Seçme
category_id_select = Select(category_id_select)
category_id_select.select_by_value("26")

# Kategori Güncelleme Alanında Yapılacak İşlem Alanını Bulma
category_transaction_select = driver.find_element(By.ID, "CategoryTransactionId")

# Kategori Güncelleme Alanında Yapılacak İşlemi Seçme
category_transaction_select = Select(category_transaction_select)
category_transaction_select.select_by_value("0")

try:
    #Kaydet Butonunu Bulma
    save_button = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, 'btn-primary'))
    )

    #Kaydet Butonununa Tıklama
    driver.execute_script("arguments[0].click();", save_button)
except Exception as e:
    print(f"Hata: {e}")


#endregion

#region Dev İndirimler Serbest Alanı Dev İndirimlerden Çıkarma

# Giriş yaptıktan sonra belirtilen sayfaya git
desired_page_url = "https://task.haydigiy.com/admin/product/bulkedit/"
driver.get(desired_page_url)

#Kategori Dahil Alan
category_select = Select(driver.find_element("id", "SearchInCategoryIds"))

#Kategori ID'si
category_select.select_by_value("529")

#Seçiniz Kısmının Tikini Kaldırma
all_remove_buttons = driver.find_elements(By.XPATH, "//span[@class='select2-selection__choice__remove']")
if len(all_remove_buttons) > 1:
    second_remove_button = all_remove_buttons[1]
    second_remove_button.click()
else:
    pass

#Ara Butonuna Tıklama
search_button = driver.find_element(By.ID, "search-products")
search_button.click()

# Sayfanın en sonuna git
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

# Kategori Güncelleme Tikine Tıklama
checkbox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "Category_Update")))
driver.execute_script("arguments[0].click();", checkbox)

# Kategori Güncelleme Alanını Bulma
category_id_select = driver.find_element(By.ID, "CategoryId")

# Kategori Güncelleme Alanında Kategori ID'si Seçme
category_id_select = Select(category_id_select)
category_id_select.select_by_value("374")

# Kategori Güncelleme Alanında Yapılacak İşlem Alanını Bulma
category_transaction_select = driver.find_element(By.ID, "CategoryTransactionId")

# Kategori Güncelleme Alanında Yapılacak İşlemi Seçme
category_transaction_select = Select(category_transaction_select)
category_transaction_select.select_by_value("1")

try:
    #Kaydet Butonunu Bulma
    save_button = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, 'btn-primary'))
    )

    #Kaydet Butonununa Tıklama
    driver.execute_script("arguments[0].click();", save_button)
except Exception as e:
    print(f"Hata: {e}")

#endregion

#region Dev İndirimler Serbest Alanı Dev İndirimlere Ekleme

# Giriş yaptıktan sonra belirtilen sayfaya git
desired_page_url = "https://task.haydigiy.com/admin/product/bulkedit/"
driver.get(desired_page_url)

#Kategori Dahil Alan
category_select = Select(driver.find_element("id", "SearchInCategoryIds"))

#Kategori ID'si
category_select.select_by_value("529")

#Seçiniz Kısmının Tikini Kaldırma
all_remove_buttons = driver.find_elements(By.XPATH, "//span[@class='select2-selection__choice__remove']")
if len(all_remove_buttons) > 1:
    second_remove_button = all_remove_buttons[1]
    second_remove_button.click()
else:
    pass

#Ara Butonuna Tıklama
search_button = driver.find_element(By.ID, "search-products")
search_button.click()

# Sayfanın en sonuna git
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

# Kategori Güncelleme Tikine Tıklama
checkbox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "Category_Update")))
driver.execute_script("arguments[0].click();", checkbox)

# Kategori Güncelleme Alanını Bulma
category_id_select = driver.find_element(By.ID, "CategoryId")

# Kategori Güncelleme Alanında Kategori ID'si Seçme
category_id_select = Select(category_id_select)
category_id_select.select_by_value("374")

# Kategori Güncelleme Alanında Yapılacak İşlem Alanını Bulma
category_transaction_select = driver.find_element(By.ID, "CategoryTransactionId")

# Kategori Güncelleme Alanında Yapılacak İşlemi Seçme
category_transaction_select = Select(category_transaction_select)
category_transaction_select.select_by_value("0")

try:
    #Kaydet Butonunu Bulma
    save_button = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, 'btn-primary'))
    )

    #Kaydet Butonununa Tıklama
    driver.execute_script("arguments[0].click();", save_button)
except Exception as e:
    print(f"Hata: {e}")

#endregion

#region Yeni Sezon Serbest Alanı Yeni Sezona Ekleme

# Giriş yaptıktan sonra belirtilen sayfaya git
desired_page_url = "https://task.haydigiy.com/admin/product/bulkedit/"
driver.get(desired_page_url)

#Kategori Dahil Alan
category_select = Select(driver.find_element("id", "SearchInCategoryIds"))

#Kategori ID'si
category_select.select_by_value("542")

#Seçiniz Kısmının Tikini Kaldırma
all_remove_buttons = driver.find_elements(By.XPATH, "//span[@class='select2-selection__choice__remove']")
if len(all_remove_buttons) > 1:
    second_remove_button = all_remove_buttons[1]
    second_remove_button.click()
else:
    pass

#Ara Butonuna Tıklama
search_button = driver.find_element(By.ID, "search-products")
search_button.click()

# Sayfanın en sonuna git
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

# Kategori Güncelleme Tikine Tıklama
checkbox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "Category_Update")))
driver.execute_script("arguments[0].click();", checkbox)

# Kategori Güncelleme Alanını Bulma
category_id_select = driver.find_element(By.ID, "CategoryId")

# Kategori Güncelleme Alanında Kategori ID'si Seçme
category_id_select = Select(category_id_select)
category_id_select.select_by_value("394")

# Kategori Güncelleme Alanında Yapılacak İşlem Alanını Bulma
category_transaction_select = driver.find_element(By.ID, "CategoryTransactionId")

# Kategori Güncelleme Alanında Yapılacak İşlemi Seçme
category_transaction_select = Select(category_transaction_select)
category_transaction_select.select_by_value("0")

try:
    #Kaydet Butonunu Bulma
    save_button = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, 'btn-primary'))
    )

    #Kaydet Butonununa Tıklama
    driver.execute_script("arguments[0].click();", save_button)
except Exception as e:
    print(f"Hata: {e}")

#endregion

#region Öne Çıkanlar Kategorisini Excelle Ürün Yükleyerek Doldurma

desired_url = "https://task.haydigiy.com/admin/importproductxls/edit/24"
driver.get(desired_url)

# Yükle Butonunu Bul
file_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name="qqfile"]')))

# CalismaAlani Excel dosyasının mevcut çalışma dizininde olduğunu varsay
file_path = os.path.join(os.getcwd(), "Öne Çıkanlar Yükleme.xlsx")

# Dosyayı seç
file_input.send_keys(file_path)


# "İşlemler" düğmesine tıkla
operations_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'btn-success')))
operations_button.click()

# Dosya yükleme işlemi bittikten sonra çalıştır butonuna tıkla
execute_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'import-product-xls-execute-confirm')))
execute_button.click()

# 10 saniye bekle
time.sleep(10)

def wait_for_element_and_click(driver, by, value, timeout=10):
    try:
        element = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))
        element.click()
        return True
    except (TimeoutException, WebDriverException) as e:
        print(f"Hata: {e}")
        return False

def wait_for_page_load(driver):
    while True:
        if driver.title:  # Tarayıcı başlığı varsa, sayfa yüklenmiş demektir
            break
        time.sleep(2)

# "Evet" butonunu tıkla
if wait_for_element_and_click(driver, By.ID, 'import-product-xls-execute'):
    # Yüklenmeyi Bekle
    wait_for_page_load(driver)

time.sleep(10)

#endregion

#region Sıralanacak Olan Kategorilerin Sırasını Bozma

# Google Sheets URL
google_sheet_url = "https://docs.google.com/spreadsheets/d/1suzb1TJyZz1xCtUxs1QOP7dHeWTKWkzDul_TaAtb1mc/gviz/tq?tqx=out:csv"

try:
    # Google Sheets'ten veriyi oku
    google_df = pd.read_csv(google_sheet_url)
    
    # "Sıralama Linkleri" sütunundaki linkleri al
    order_edit_urls = google_df['Sıralama Linkleri'].tolist()
    
except requests.exceptions.RequestException as e:
    print("Google Sheets'e erişilemiyor:", e)
except pd.errors.EmptyDataError:
    print("Google Sheets'ten veri okunamadı.")

try:
    # Her bir link için işlem yapma
    for url in order_edit_urls:
        driver.get(url)

        js_code = """
        document.getElementById("SortOptionId").value = "25";
        var event = new Event('change');
        document.getElementById("SortOptionId").dispatchEvent(event);

        document.getElementById("btnChangeSorting").click();
        document.getElementById("btnChangeSorting-action-confirmation-submit-button").click();
        """
        driver.execute_script(js_code)

        # İşlem tamamlanana kadar bekleyin (maksimum 10 saniye)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "btnChangeSorting")))

except Exception as e:
    print("Hata oluştu:", e)



#endregion

#region Yeni Gelenler Kategorisinin Sırasını Bozma


# order_edit_urls listesi
order_edit_urls = [
        "https://task.haydigiy.com/Admin/Category/Sort/394"


]

try:

    # Her bir link için işlem yapma
    for url in order_edit_urls:
        driver.get(url)

        js_code = """
        document.getElementById("btnChangeSorting").click();
        document.getElementById("btnChangeSorting-action-confirmation-submit-button").click();
        """
        driver.execute_script(js_code)

        # İşlem tamamlanana kadar bekleyin (maksimum 10 saniye)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "btnChangeSorting")))

except Exception as e:
    pass

#endregion

#region İç Giyim Ürünlerini Öne Çıkanlardan Çıkarma

# Giriş yaptıktan sonra belirtilen sayfaya git
desired_page_url = "https://task.haydigiy.com/admin/product/bulkedit/"
driver.get(desired_page_url)

#Kategori Dahil Alan
category_select = Select(driver.find_element("id", "SearchInCategoryIds"))

#Kategori ID'si
category_select.select_by_value("172")

#Seçiniz Kısmının Tikini Kaldırma
all_remove_buttons = driver.find_elements(By.XPATH, "//span[@class='select2-selection__choice__remove']")
if len(all_remove_buttons) > 1:
    second_remove_button = all_remove_buttons[1]
    second_remove_button.click()
else:
    pass

#Ara Butonuna Tıklama
search_button = driver.find_element(By.ID, "search-products")
search_button.click()

# Sayfanın en sonuna git
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

# Kategori Güncelleme Tikine Tıklama
checkbox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "Category_Update")))
driver.execute_script("arguments[0].click();", checkbox)

# Kategori Güncelleme Alanını Bulma
category_id_select = driver.find_element(By.ID, "CategoryId")

# Kategori Güncelleme Alanında Kategori ID'si Seçme
category_id_select = Select(category_id_select)
category_id_select.select_by_value("26")

# Kategori Güncelleme Alanında Yapılacak İşlem Alanını Bulma
category_transaction_select = driver.find_element(By.ID, "CategoryTransactionId")

# Kategori Güncelleme Alanında Yapılacak İşlemi Seçme
category_transaction_select = Select(category_transaction_select)
category_transaction_select.select_by_value("1")

try:
    #Kaydet Butonunu Bulma
    save_button = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, 'btn-primary'))
    )

    #Kaydet Butonununa Tıklama
    driver.execute_script("arguments[0].click();", save_button)
except Exception as e:
    print(f"Hata: {e}")

# Tarayıcıyı kapatma
driver.quit()

#endregion

#region Kategorileri Sıralama


# Global tken değişkeni
_auth_token = None

# Token alma fonksiyonu
def get_auth_token():
    global _auth_token
    if _auth_token is None:  
        login_url = "https://siparis.haydigiy.com/api/customer/login"
        login_payload = {
            "apiKey": "MypGcaEInEOTzuYQydgDHQ",
            "secretKey": "jRqliBLDPke76YhL_WL5qg",
            "emailOrPhone": "mustafa_kod@haydigiy.com",
            "password": "123456"
        }
        login_headers = {
            "Content-Type": "application/json"
        }

        response = requests.post(login_url, json=login_payload, headers=login_headers)
        if response.status_code == 200:
            _auth_token = response.json().get("data", {}).get("token")
            if not _auth_token:
                raise Exception("TOKEN ALINAMADI")
        else:
            raise Exception(f"GİRİŞ BAŞARISIZ: {response.text}")
    return _auth_token

# Token alma işlemi
token = get_auth_token()
df = pd.read_excel("Son Liste.xlsx")
conn = http.client.HTTPSConnection("siparis.haydigiy.com")


for index, row in tqdm(df.iterrows(), total=df.shape[0], desc="Normal Kategoriler Sıralanıyor"):

    category_id = row['Kategori ID']
    display_order = row['Numara']
    product_id = row['ID']

    product_id = str(row['ID']).replace(".0", "")

    payload = json.dumps({
        "CategoryId": int(category_id),  
        "IsFeaturedProduct": False, 
        "DisplayOrder": int(display_order)
    })

    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {token}',
        'Cookie': '.Application.Customer=64684894-1b54-488d-bd59-76b94842df65'
    }


    conn.request("PUT", f"/adminapi/product/product-categories?productId={product_id}", payload, headers)
    res = conn.getresponse()
    data = res.read()


#endregion

#region Cloudflare Önbellek Temizleme

def send_links_to_api(links):
    try:
        # Cloudflare API'ye istek gönder
        cf_url = "https://api.cloudflare.com/client/v4/zones/469d52dc478eb1a2e1864dc0b3f548ac/purge_cache"
        headers = {
            "Content-Type": "application/json",
            "X-Auth-Email": "erkan@haydigiy.com",
            "X-Auth-Key": "c45a4d56745100a8568a2e9e7a00948f23b4e"
        }
        data = {
            "files": links
        }

        # Gönderilen linkleri yazdırma
        print("API'ye gönderilen linkler:")
        for link in links:
            print(link)

        cf_response = requests.post(cf_url, headers=headers, json=data)
        cf_response.raise_for_status()
        print("Linkler başarıyla gönderildi!")

    except requests.exceptions.HTTPError as e:
        print(f"API isteği başarısız oldu: {e}, başa dön")

def fetch_and_send_links():
    links = [
        "https://www.haydigiy.com/aksesuar/",
        "https://www.haydigiy.com/kadin-canta/",
        "https://www.haydigiy.com/kadin-gozluk/",
        "https://www.haydigiy.com/kolye-kupe-bileklik/",
        "https://www.haydigiy.com/kadin-sapka/",
        "https://www.haydigiy.com/parfum/",
        "https://www.haydigiy.com/sal/",
        "https://www.haydigiy.com/saat-3/",
        "https://www.haydigiy.com/kadin-kemer/",
        "https://www.haydigiy.com/indirimli-urunler/",
        "https://www.haydigiy.com/ust-giyim/",
        "https://www.haydigiy.com/kadin-gomlek/",
        "https://www.haydigiy.com/kadin-hirka-2/",
        "https://www.haydigiy.com/kadin-sweat/",
        "https://www.haydigiy.com/kadin-tisort/",
        "https://www.haydigiy.com/tunik/",
        "https://www.haydigiy.com/kadin-yelek/",
        "https://www.haydigiy.com/kadin-atlet/",
        "https://www.haydigiy.com/bluz-body-triko/",
        "https://www.haydigiy.com/ceket/",
        "https://www.haydigiy.com/kot-ceket/",
        "https://www.haydigiy.com/kadin-ceket-mont-2/",
        "https://www.haydigiy.com/tum-urunler-2/",
        "https://www.haydigiy.com/imt-olmayan/",
        "https://www.haydigiy.com/kadin-buyuk-beden/",
        "https://www.haydigiy.com/ic-giyim-tum-urunler/",
        "https://www.haydigiy.com/fantezi/",
        "https://www.haydigiy.com/fantezi-ic-giyim/",
        "https://www.haydigiy.com/jartiyer/",
        "https://www.haydigiy.com/fantezi-gecelik/",
        "https://www.haydigiy.com/fantezi-kostum/",
        "https://www.haydigiy.com/kadin/",
        "https://www.haydigiy.com/kadin-gecelik/",
        "https://www.haydigiy.com/kadin-kulot/",
        "https://www.haydigiy.com/sutyen-takimlari/",
        "https://www.haydigiy.com/penti/",
        "https://www.haydigiy.com/kulotlu-corap/",
        "https://www.haydigiy.com/diz-alti-coraplar/",
        "https://www.haydigiy.com/soket-corap/",
        "https://www.haydigiy.com/plaj-2/",
        "https://www.haydigiy.com/mayo/",
        "https://www.haydigiy.com/plaj-elbisesi/",
        "https://www.haydigiy.com/bikini/",
        "https://www.haydigiy.com/tesettur-mayo/",
        "https://www.haydigiy.com/hasema/",
        "https://www.haydigiy.com/fiyata-hamle/",
        "https://www.haydigiy.com/haric-tutulacaklar/",
        "https://www.haydigiy.com/stoklu-ic-giyim/",
        "https://www.haydigiy.com/reklam-kategorisi/",
        "https://www.haydigiy.com/yeni-sezon/",
        "https://www.haydigiy.com/ic-giyim-model-cekimi/",
        "https://www.haydigiy.com/mustafa-urun-listesi-haric-tutulacaklar/",
        "https://www.haydigiy.com/unisex-urunler/",
        "https://www.haydigiy.com/fiyata-hamle-2/",
        "https://www.haydigiy.com/test-kategorisi/",
        "https://www.haydigiy.com/mustafa-ozel/",
        "https://www.haydigiy.com/sezon-sonu/",
        "https://www.haydigiy.com/cocuk-2/",
        "https://www.haydigiy.com/cocuk-ayakkabi/",
        "https://www.haydigiy.com/cocuk-terlik/",
        "https://www.haydigiy.com/cocuk-canta/",
        "https://www.haydigiy.com/cocuk-takim-2/",
        "https://www.haydigiy.com/cocuk-kot-pantolon/",
        "https://www.haydigiy.com/cocuk-ust-giyim/",
        "https://www.haydigiy.com/cocuk-elbise/",
        "https://www.haydigiy.com/cocuk-pijama-takimi/",
        "https://www.haydigiy.com/kot-pantolon-firsatlari/",
        "https://www.haydigiy.com/erkek-3/",
        "https://www.haydigiy.com/erkek-esofman-alti-2/",
        "https://www.haydigiy.com/erkek-ceket-mont-kaban/",
        "https://www.haydigiy.com/erkek-ceket-mont-kaban-2/",
        "https://www.haydigiy.com/erkek-tisort-2/",
        "https://www.haydigiy.com/erkek-ayakkabi-2/",
        "https://www.haydigiy.com/goruntulenme-urunu/",
        "https://www.haydigiy.com/alis-fiyati-kontrol-edildi/",
        "https://www.haydigiy.com/cocuk-ana-sayfa/",
        "https://www.haydigiy.com/ic-giyim-ana-sayfa/",
        "https://www.haydigiy.com/masa-kombini-olan-ic-giyim-urunleri/",
        "https://www.haydigiy.com/silinecek-urunler-2/",
        "https://www.haydigiy.com/bayram-senligi/",
        "https://www.haydigiy.com/arama-terimleri-girildi/",
        "https://www.haydigiy.com/ifondi-siralama-ilk-10-urun/",
        "https://www.haydigiy.com/urun-ozelligi-girilen-urunler/",
        "https://www.haydigiy.com/hatali-urunler-2/",
        "https://www.haydigiy.com/tesettur-3/",
        "https://www.haydigiy.com/tesettur-ana-sayfa/",
        "https://www.haydigiy.com/ic-giyim-test/",
        "https://www.haydigiy.com/senin-icin-sectiklerimiz/",
        "https://www.haydigiy.com/1-tl-ye-urun/",
        "https://www.haydigiy.com/1-tl/",
        "https://www.haydigiy.com/5-olarak-kalacak-urunler/",
        "https://www.haydigiy.com/etto-park/",
        "https://www.haydigiy.com/liste-fiyatindan-olacak-urunler/",
        "https://www.haydigiy.com/test-kategori/",
        "https://www.haydigiy.com/stogu-0-olan-urunler/",
        "https://www.haydigiy.com/gecici-kategori/",
        "https://www.haydigiy.com/maliyetine-satislar-firsati/",
        "https://www.haydigiy.com/one-cikanlar-serbest-alan/",
        "https://www.haydigiy.com/dev-indirimler-serbest-alan-2/",
        "https://www.haydigiy.com/kategori-test/",
        "https://www.haydigiy.com/mustafa-tukenen-urunler/",
        "https://www.haydigiy.com/dis-giyim-2/",
        "https://www.haydigiy.com/askidaki-urunler/",
        "https://www.haydigiy.com/indirimli-cocuk-urunleri/",
        "https://www.haydigiy.com/alt-giyim/",
        "https://www.haydigiy.com/etek/",
        "https://www.haydigiy.com/kot-pantolon/",
        "https://www.haydigiy.com/kadin-pantolon/",
        "https://www.haydigiy.com/kadin-tayt/",
        "https://www.haydigiy.com/sort/",
        "https://www.haydigiy.com/kadin-ayakkabi/",
        "https://www.haydigiy.com/kadin-babet-sandalet-terlik/",
        "https://www.haydigiy.com/gunluk-ayakkabi/",
        "https://www.haydigiy.com/topuklu-ayakkabi/",
        "https://www.haydigiy.com/one-cikanlar/",
        "https://www.haydigiy.com/elbise-tulum/",
        "https://www.haydigiy.com/elbise/",
        "https://www.haydigiy.com/kadin-tulum-takim/",
        "https://www.haydigiy.com/kadin-esofman-pijama/",
        "https://www.haydigiy.com/kadin-esofman-alti/",
        "https://www.haydigiy.com/kadin-esofman-takimi/",
        "https://www.haydigiy.com/kadin-pijama-takimi/",
        "https://www.haydigiy.com/imt/",
        "https://www.haydigiy.com/tek-kalanlar/",
        "https://www.haydigiy.com/yeni-gelenler/",
        "https://www.haydigiy.com/cok-satanlar/",
        "https://www.haydigiy.com/tekrar-stokta/",
        "https://www.haydigiy.com/75-tl-alti-urunler/",
        "https://www.haydigiy.com/davet-kiyafetleri/",
        "https://www.haydigiy.com/aksesuar",
        "https://www.haydigiy.com/kadin-canta",
        "https://www.haydigiy.com/kadin-gozluk",
        "https://www.haydigiy.com/kolye-kupe-bileklik",
        "https://www.haydigiy.com/kadin-sapka",
        "https://www.haydigiy.com/parfum",
        "https://www.haydigiy.com/sal",
        "https://www.haydigiy.com/saat-3",
        "https://www.haydigiy.com/kadin-kemer",
        "https://www.haydigiy.com/indirimli-urunler",
        "https://www.haydigiy.com/ust-giyim",
        "https://www.haydigiy.com/kadin-gomlek",
        "https://www.haydigiy.com/kadin-hirka-2",
        "https://www.haydigiy.com/kadin-sweat",
        "https://www.haydigiy.com/kadin-tisort",
        "https://www.haydigiy.com/tunik",
        "https://www.haydigiy.com/kadin-yelek",
        "https://www.haydigiy.com/kadin-atlet",
        "https://www.haydigiy.com/bluz-body-triko",
        "https://www.haydigiy.com/ceket",
        "https://www.haydigiy.com/kot-ceket",
        "https://www.haydigiy.com/kadin-ceket-mont-2",
        "https://www.haydigiy.com/tum-urunler-2",
        "https://www.haydigiy.com/imt-olmayan",
        "https://www.haydigiy.com/kadin-buyuk-beden",
        "https://www.haydigiy.com/ic-giyim-tum-urunler",
        "https://www.haydigiy.com/fantezi",
        "https://www.haydigiy.com/fantezi-ic-giyim",
        "https://www.haydigiy.com/jartiyer",
        "https://www.haydigiy.com/fantezi-gecelik",
        "https://www.haydigiy.com/fantezi-kostum",
        "https://www.haydigiy.com/kadin",
        "https://www.haydigiy.com/kadin-gecelik",
        "https://www.haydigiy.com/kadin-kulot",
        "https://www.haydigiy.com/sutyen-takimlari",
        "https://www.haydigiy.com/penti",
        "https://www.haydigiy.com/kulotlu-corap",
        "https://www.haydigiy.com/diz-alti-coraplar",
        "https://www.haydigiy.com/soket-corap",
        "https://www.haydigiy.com/plaj-2",
        "https://www.haydigiy.com/mayo",
        "https://www.haydigiy.com/plaj-elbisesi",
        "https://www.haydigiy.com/bikini",
        "https://www.haydigiy.com/tesettur-mayo",
        "https://www.haydigiy.com/hasema",
        "https://www.haydigiy.com/fiyata-hamle",
        "https://www.haydigiy.com/haric-tutulacaklar",
        "https://www.haydigiy.com/stoklu-ic-giyim",
        "https://www.haydigiy.com/reklam-kategorisi",
        "https://www.haydigiy.com/yeni-sezon",
        "https://www.haydigiy.com/ic-giyim-model-cekimi",
        "https://www.haydigiy.com/mustafa-urun-listesi-haric-tutulacaklar",
        "https://www.haydigiy.com/unisex-urunler",
        "https://www.haydigiy.com/fiyata-hamle-2",
        "https://www.haydigiy.com/test-kategorisi",
        "https://www.haydigiy.com/mustafa-ozel",
        "https://www.haydigiy.com/sezon-sonu",
        "https://www.haydigiy.com/cocuk-2",
        "https://www.haydigiy.com/cocuk-ayakkabi",
        "https://www.haydigiy.com/cocuk-terlik",
        "https://www.haydigiy.com/cocuk-canta",
        "https://www.haydigiy.com/cocuk-takim-2",
        "https://www.haydigiy.com/cocuk-kot-pantolon",
        "https://www.haydigiy.com/cocuk-ust-giyim",
        "https://www.haydigiy.com/cocuk-elbise",
        "https://www.haydigiy.com/cocuk-pijama-takimi",
        "https://www.haydigiy.com/kot-pantolon-firsatlari",
        "https://www.haydigiy.com/erkek-3",
        "https://www.haydigiy.com/erkek-esofman-alti-2",
        "https://www.haydigiy.com/erkek-ceket-mont-kaban",
        "https://www.haydigiy.com/erkek-ceket-mont-kaban-2",
        "https://www.haydigiy.com/erkek-tisort-2",
        "https://www.haydigiy.com/erkek-ayakkabi-2",
        "https://www.haydigiy.com/goruntulenme-urunu",
        "https://www.haydigiy.com/alis-fiyati-kontrol-edildi",
        "https://www.haydigiy.com/cocuk-ana-sayfa",
        "https://www.haydigiy.com/ic-giyim-ana-sayfa",
        "https://www.haydigiy.com/masa-kombini-olan-ic-giyim-urunleri",
        "https://www.haydigiy.com/silinecek-urunler-2",
        "https://www.haydigiy.com/bayram-senligi",
        "https://www.haydigiy.com/arama-terimleri-girildi",
        "https://www.haydigiy.com/ifondi-siralama-ilk-10-urun",
        "https://www.haydigiy.com/urun-ozelligi-girilen-urunler",
        "https://www.haydigiy.com/hatali-urunler-2",
        "https://www.haydigiy.com/tesettur-3",
        "https://www.haydigiy.com/tesettur-ana-sayfa",
        "https://www.haydigiy.com/ic-giyim-test",
        "https://www.haydigiy.com/senin-icin-sectiklerimiz",
        "https://www.haydigiy.com/1-tl-ye-urun",
        "https://www.haydigiy.com/1-tl",
        "https://www.haydigiy.com/5-olarak-kalacak-urunler",
        "https://www.haydigiy.com/etto-park",
        "https://www.haydigiy.com/liste-fiyatindan-olacak-urunler",
        "https://www.haydigiy.com/test-kategori",
        "https://www.haydigiy.com/stogu-0-olan-urunler",
        "https://www.haydigiy.com/gecici-kategori",
        "https://www.haydigiy.com/maliyetine-satislar-firsati",
        "https://www.haydigiy.com/one-cikanlar-serbest-alan",
        "https://www.haydigiy.com/dev-indirimler-serbest-alan-2",
        "https://www.haydigiy.com/kategori-test",
        "https://www.haydigiy.com/mustafa-tukenen-urunler",
        "https://www.haydigiy.com/dis-giyim-2",
        "https://www.haydigiy.com/askidaki-urunler",
        "https://www.haydigiy.com/indirimli-cocuk-urunleri",
        "https://www.haydigiy.com/alt-giyim",
        "https://www.haydigiy.com/etek",
        "https://www.haydigiy.com/kot-pantolon",
        "https://www.haydigiy.com/kadin-pantolon",
        "https://www.haydigiy.com/kadin-tayt",
        "https://www.haydigiy.com/sort",
        "https://www.haydigiy.com/kadin-ayakkabi",
        "https://www.haydigiy.com/kadin-babet-sandalet-terlik",
        "https://www.haydigiy.com/gunluk-ayakkabi",
        "https://www.haydigiy.com/topuklu-ayakkabi",
        "https://www.haydigiy.com/one-cikanlar",
        "https://www.haydigiy.com/elbise-tulum",
        "https://www.haydigiy.com/elbise",
        "https://www.haydigiy.com/kadin-tulum-takim",
        "https://www.haydigiy.com/kadin-esofman-pijama",
        "https://www.haydigiy.com/kadin-esofman-alti",
        "https://www.haydigiy.com/kadin-esofman-takimi",
        "https://www.haydigiy.com/kadin-pijama-takimi",
        "https://www.haydigiy.com/imt",
        "https://www.haydigiy.com/tek-kalanlar",
        "https://www.haydigiy.com/yeni-gelenler",
        "https://www.haydigiy.com/cok-satanlar",
        "https://www.haydigiy.com/tekrar-stokta",
        "https://www.haydigiy.com/75-tl-alti-urunler",
        "https://www.haydigiy.com/davet-kiyafetleri"

    ]

    # Linkleri 15'li gruplara ayır
    chunk_size = 15
    for i in range(0, len(links), chunk_size):
        chunk = links[i:i + chunk_size]
        send_links_to_api(chunk)

if __name__ == "__main__":
    fetch_and_send_links()

#endregion



#region Kategorilerdeki Ürün Adedini Haydigiy Online'a Gönderme

# Hedef URL
url = 'https://task.haydigiy.com/iletisim/'

# URL'ye istek gönder
response = requests.get(url)

# İçeriği parse et
soup = BeautifulSoup(response.content, 'html.parser')

# Linkleri ve isimleri saklamak için bir liste oluştur
all_links = []

# Script içeriğini bul
script = soup.find('script', text=lambda t: t and 'header-menu2' in t)

# Script içeriğinden metni al
script_content = script.string

# Menüyü oluşturacak HTML parçalarını belirle
start_index = script_content.index('<div class=header-menu2>')
end_index = script_content.index('</div>', start_index) + len('</div>')
menu_html = script_content[start_index:end_index]

# Menü HTML'ini parse et
menu_soup = BeautifulSoup(menu_html, 'html.parser')

# Linkleri ve isimleri bul
links = menu_soup.find_all('a')

# Tüm linkleri sakla
for link in links:
    href = link.get('href')
    name = link.text.strip()
    all_links.append((href, name))

# "header-menu" sınıfındaki linkleri bul
header_menu_div = soup.find('div', class_='header-menu')

# Tüm linkleri sakla
header_links = header_menu_div.find_all('a')

for link in header_links:
    href = link.get('href')
    name = link.text.strip()
    all_links.append((href, name))

def fetch_product_count(href, name):
    # Tam URL oluştur
    full_url = f'https://task.haydigiy.com{href}'
    link_response = requests.get(full_url)

    # Sayfanın kaynak HTML'ini al
    html_content = link_response.text

    # JavaScript içindeki categoryId'yi ayıklamak için bir düzenli ifade (regex) kullanıyoruz
    match = re.search(r'var categoryId\s*=\s*(\d+);', html_content)

    category_id = match.group(1)  # İlk eşleşmeyi al
    total_product_count = 0  # Toplam ürün sayısını sıfırla
    page_number = 1  # Sayfa numarasını başlat

    while True:
        # AJAX URL'sini oluştur
        ajax_url = f'https://task.haydigiy.com/Catalog/AjaxCategory/?categoryId={category_id}&pageNumber={page_number}&pageSize=12'
        ajax_response = requests.get(ajax_url)

        # Sayfanın kaynak HTML'ini al
        ajax_html_content = ajax_response.text
        ajax_soup = BeautifulSoup(ajax_html_content, 'html.parser')

        # product-item sınıfındaki öğeleri say
        product_items = ajax_soup.find_all(class_='product-item')
        product_count = len(product_items)

        if product_count == 0:
            break  # Daha fazla sayfaya istek göndermeyi durdur

        total_product_count += product_count  # Toplama ekle
        page_number += 1  # Sonraki sayfaya geç

    return category_id, total_product_count

# 10'arlı paralel istek gönder
with ThreadPoolExecutor(max_workers=10) as executor:
    future_to_link = {executor.submit(fetch_product_count, href, name): (href, name) for href, name in all_links}

    for future in as_completed(future_to_link):
        href, name = future_to_link[future]
        category_id, total_product_count = future.result()

        # Verileri GET isteği olarak gönder
        get_url = f"https://haydigiy.online/haydigiy/numberofitems.php?url={href}&name={name}&categoryid={category_id}&totalproducts={total_product_count}"
        requests.get(get_url)


#endregion

#region Kategorilerin İçindeki Kategorileri Tespit Edip Haydigiy Online'a Gönderme

# İstek göndereceğimiz ana URL
url = 'https://task.haydigiy.com/iletisim/'

# HTTP isteği gönder
response = requests.get(url)

# HTML içeriğini parse et
soup = BeautifulSoup(response.content, 'html.parser')

# class="header-menu-container" sınıfını bul
header_menu_container = soup.find('div', class_='header-menu-container')

# class="one-level" olan tüm <li> elemanlarını bul
one_level_items = header_menu_container.find_all('li', class_='one-level')

# Her bir <li> elemanı için href'i al ve o linke istek gönder
for item in one_level_items:
    # <a> etiketini bul
    a_tag = item.find('a')
    
    # href değerini al
    href = a_tag['href']
    full_url = f"https://task.haydigiy.com{href}"  # Tam URL oluştur
    
    # Alt sayfaya istek gönder
    sub_response = requests.get(full_url)
    
    # Alt sayfanın HTML içeriğini parse et
    sub_soup = BeautifulSoup(sub_response.content, 'html.parser')

    # class="filter-box notfiltered-items" kısmını bul
    filter_box = sub_soup.find('div', class_='filter-box notfiltered-items')

    # span elemanlarını bul
    spans = filter_box.find_all('span')

    # Bulunan span elemanlarının metinlerini al ve "Kategori Seçiniz" olmayanları topla
    categories = [span.get_text().strip() for span in spans if span.get_text().strip() != "Kategori Seçiniz"]

    # Kategorileri virgülle birleştir
    category_str = ', '.join(categories)

    # İstek gönderilecek URL
    target_url = f"https://haydigiy.online/haydigiy/numberofcategory.php?link={full_url}&category={category_str}"

    # Veriyi göndermek için istek yap
    requests.get(target_url)

#endregion
